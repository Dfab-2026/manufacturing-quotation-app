from __future__ import annotations
from app.extraction.pipeline import analyze_pdf_with_ai
from app.db import (
    append_extraction_record,
    append_review_record,
    database_stats,
    dataset_counts_fast,
    db_ping,
    init_database,
    latest_review_by_hash,
    load_store,
    save_store,
    training_samples_for_export,
    upsert_training_sample,
)

from datetime import datetime, timezone
from hashlib import sha256
from io import BytesIO
from pathlib import Path
from typing import Literal
import base64
import json
import os
import math
import re
import uuid
import zipfile

import pymupdf as fitz

from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from pydantic import BaseModel, Field
from pypdf import PdfReader
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

APP_DIR = Path(__file__).resolve().parent
LEGACY_DATA_DIR = APP_DIR.parent / "data"


def now() -> str:
    return datetime.now(timezone.utc).isoformat()


_HOT_CACHE: dict[str, object] = {}


def load(name: str, default):
    if name in {"rates", "settings"} and name in _HOT_CACHE:
        return _HOT_CACHE[name]

    value = load_store(name, default)

    if name in {"rates", "settings"}:
        _HOT_CACHE[name] = value

    return value


def save(name: str, data) -> None:
    save_store(name, data)

    if name in {"rates", "settings"}:
        _HOT_CACHE[name] = data

def defaults_settings():
    return {
        "company_name": "AI Manufacturing Quotation",
        "currency": "INR",
        "material_wastage_pct": 8.0,
        "overhead_pct": 12.0,
        "markup_pct": 15.0,
        "auto_dataset_capture": True,
        "learn_from_corrections": True,
        "training_batch_threshold": 25,
        "critical_medium_threshold": 40,
        "critical_high_threshold": 70,
    }


def _rate(
    rid: str,
    category: str,
    name: str,
    grade: str,
    unit: str,
    price: float,
    critical: int,
    notes: str,
):
    return {
        "id": rid,
        "category": category,
        "name": name,
        "grade": grade,
        "unit": unit,
        "price": price,
        "critical_score": critical,
        "active": True,
        "notes": notes,
        "updated_at": now(),
    }


def default_rate_items():
    # These are editable starter values, not live market prices.
    return [
        _rate("MAT-SS-201", "MATERIAL", "Stainless Steel", "AISI 201", "kg", 210, 88, "Starter rate - replace with supplier rate"),
        _rate("MAT-SS-202", "MATERIAL", "Stainless Steel", "AISI 202", "kg", 225, 88, "Starter rate - replace with supplier rate"),
        _rate("MAT-SS-304", "MATERIAL", "Stainless Steel", "AISI 304", "kg", 280, 92, "Starter rate - replace with supplier rate"),
        _rate("MAT-SS-304L", "MATERIAL", "Stainless Steel", "AISI 304L", "kg", 295, 92, "Starter rate - replace with supplier rate"),
        _rate("MAT-SS-316", "MATERIAL", "Stainless Steel", "AISI 316", "kg", 380, 94, "Starter rate - replace with supplier rate"),
        _rate("MAT-SS-316L", "MATERIAL", "Stainless Steel", "AISI 316L", "kg", 400, 95, "Starter rate - replace with supplier rate"),
        _rate("MAT-SS-321", "MATERIAL", "Stainless Steel", "AISI 321", "kg", 410, 93, "Starter rate - replace with supplier rate"),
        _rate("MAT-SS-430", "MATERIAL", "Stainless Steel", "AISI 430", "kg", 190, 85, "Starter rate - replace with supplier rate"),
        _rate("MAT-MS-E250", "MATERIAL", "Mild Steel", "IS 2062 E250", "kg", 70, 82, "Starter rate - replace with supplier rate"),
        _rate("MAT-MS-E350", "MATERIAL", "Mild Steel", "IS 2062 E350", "kg", 78, 82, "Starter rate - replace with supplier rate"),
        _rate("MAT-CS-A36", "MATERIAL", "Carbon Steel", "ASTM A36", "kg", 75, 80, "Starter rate - replace with supplier rate"),
        _rate("MAT-CS-S355J0", "MATERIAL", "Carbon / Structural Steel", "S355J0 / 1.0553", "kg", 85, 95, "Starter rate - edit in Rate Master"),

        _rate("MAT-AL-5052", "MATERIAL", "Aluminium", "5052-H32", "kg", 310, 88, "Starter rate - replace with supplier rate"),
        _rate("MAT-AL-6061", "MATERIAL", "Aluminium", "6061-T6", "kg", 340, 88, "Starter rate - replace with supplier rate"),
        _rate("MAT-AL-6082", "MATERIAL", "Aluminium", "6082-T6", "kg", 360, 88, "Starter rate - replace with supplier rate"),
        _rate("MAT-GI-DX51", "MATERIAL", "Galvanized Steel", "DX51D+Z", "kg", 95, 84, "Starter rate - replace with supplier rate"),
        _rate("MAT-CU-C110", "MATERIAL", "Copper", "C110", "kg", 780, 96, "Starter rate - replace with supplier rate"),
        _rate("MAT-BR-C260", "MATERIAL", "Brass", "C260", "kg", 520, 92, "Starter rate - replace with supplier rate"),
        _rate("PROC-LASER", "PROCESS", "Laser Cutting", "", "job", 450, 75, "Machine/setup rate"),
        _rate("PROC-BEND", "PROCESS", "Press Brake Forming", "", "job", 350, 70, "Bending/forming rate"),
        _rate("PROC-TIG", "PROCESS", "TIG Welding", "", "m", 250, 85, "Weld rate per metre"),
        _rate("PROC-MIG", "PROCESS", "MIG Welding", "", "m", 180, 80, "Weld rate per metre"),
        _rate("PROC-STUD", "PROCESS", "Stud Welding", "", "stud", 50, 65, "Rate per stud"),
        _rate("PROC-GRIND", "PROCESS", "Grinding & Flush", "", "job", 350, 70, "Finishing rate"),
        _rate("PROC-DEBURR", "PROCESS", "Deburring", "", "job", 150, 55, "Edge cleanup rate"),
        _rate("PROC-DRILL", "PROCESS", "Drilling", "", "hole", 15, 55, "Rate per hole"),
        _rate("PROC-POLISH", "PROCESS", "Polishing", "", "m2", 600, 75, "Surface finishing rate"),
        _rate("PROC-PASS", "PROCESS", "Passivation", "", "job", 500, 70, "Passivation allowance"),
        _rate("PROC-COAT", "PROCESS", "Powder Coating", "", "m2", 300, 75, "Coating rate"),
        _rate("PROC-QC", "PROCESS", "Inspection & Handling", "", "job", 150, 45, "Inspection allowance"),
        _rate("PROC-WELD", "PROCESS", "General / Tack Welding", "", "job", 300, 80, "Starter rate - edit in Rate Master"),
        _rate("PROC-SAW", "PROCESS", "Saw / Raw Stock Cutting", "", "job", 200, 70, "Starter rate - edit in Rate Master"),
        _rate("PROC-TURN", "PROCESS", "CNC Turning", "", "hr", 650, 90, "Starter hourly machine rate - edit in Rate Master"),
        _rate("PROC-BORE", "PROCESS", "Drilling / Boring", "", "job", 350, 85, "Starter rate - edit in Rate Master"),
        _rate("PROC-THREAD", "PROCESS", "Threading / Tapping", "", "job", 250, 90, "Starter rate - edit in Rate Master"),
        _rate("PROC-CHAMFER", "PROCESS", "Chamfering", "", "job", 120, 65, "Starter rate - edit in Rate Master"),
        _rate("PROC-MACHINE", "PROCESS", "General Machining", "", "hr", 700, 90, "Starter hourly machine rate - edit in Rate Master"),
        _rate("PROC-MILL", "PROCESS", "CNC Milling", "", "hr", 800, 90, "Starter hourly machine rate - edit in Rate Master"),
        _rate("PROC-MANUAL-MILL", "PROCESS", "Manual Milling", "", "hr", 550, 80, "Starter hourly machine rate - edit in Rate Master"),
        _rate("PROC-SHEAR", "PROCESS", "Shearing", "", "job", 250, 65, "Starter job rate - edit in Rate Master"),
        _rate("PROC-PLASMA", "PROCESS", "Plasma Cutting", "", "job", 500, 75, "Starter job rate - edit in Rate Master"),
        _rate("PROC-WATERJET", "PROCESS", "Waterjet Cutting", "", "job", 800, 75, "Starter job rate - edit in Rate Master"),
        _rate("PROC-HAND-GRIND", "PROCESS", "Hand Grinding / Cut-off", "", "hr", 350, 70, "Starter hourly labour/machine allowance - edit in Rate Master"),

        _rate("LAB-FAB", "LABOUR", "Fabricator", "", "hr", 350, 65, "Skilled fabricator labour"),
        _rate("LAB-TIG", "LABOUR", "TIG Welder", "", "hr", 450, 75, "TIG welder labour"),
        _rate("LAB-MIG", "LABOUR", "MIG Welder", "", "hr", 400, 70, "MIG welder labour"),
        _rate("LAB-FINISH", "LABOUR", "Finishing Operator", "", "hr", 300, 55, "Grinding/polishing labour"),
        _rate("LAB-QC", "LABOUR", "QC / Handling", "", "hr", 250, 45, "QC and handling labour"),
        _rate("LAB-MACH", "LABOUR", "Machinist", "", "hr", 500, 78, "Starter rate - edit in Rate Master"),
        _rate("LAB-MACHINE", "LABOUR", "Machine Operator", "", "hr", 350, 65, "Starter rate - edit in Rate Master"),
        _rate("LAB-WELD", "LABOUR", "Welder / Fabricator", "", "hr", 400, 75, "Starter rate - edit in Rate Master"),

        _rate("OTHER-PACK", "OTHER", "Packing", "", "job", 250, 40, "Packing allowance"),
        _rate("OTHER-CONS", "OTHER", "Welding Consumables", "", "job", 200, 60, "Consumables allowance"),
        _rate("COMM-WASTE", "COMMERCIAL", "Material Wastage", "", "%", 8, 60, "Applied to material cost"),
        _rate("COMM-OH", "COMMERCIAL", "Factory Overhead", "", "%", 12, 65, "Applied to direct cost + material wastage"),
        _rate("COMM-MARKUP", "COMMERCIAL", "Profit / Markup", "", "%", 15, 80, "Applied to manufacturing cost"),
    ]


def _repair_default_rates(existing_rates: list[dict] | None) -> tuple[list[dict], int]:
    """
    Ensure every built-in starter Rate Master row exists.

    Rules:
    - Never overwrite a positive engineer-entered rate.
    - Re-add a built-in row if it was missing.
    - Fill a zero/blank built-in rate from the starter default.
    - Fill blank metadata only.
    - Keep all custom/user-created Rate Master rows.
    """
    defaults = default_rate_items()
    rows = [
        dict(row)
        for row in (existing_rates or [])
        if isinstance(row, dict)
    ]

    by_id = {
        str(row.get("id") or ""): row
        for row in rows
        if row.get("id")
    }

    changed = 0

    for default in defaults:
        default_id = str(default["id"])
        row = by_id.get(default_id)

        if row is None:
            clone = dict(default)
            rows.append(clone)
            by_id[default_id] = clone
            changed += 1
            continue

        # Preserve approved/manual positive prices. Only repair missing/zero values.
        if float(row.get("price", 0) or 0) <= 0 and float(default.get("price", 0) or 0) > 0:
            row["price"] = default["price"]
            row["updated_at"] = now()
            changed += 1

        for key in ("category", "name", "grade", "unit"):
            if not str(row.get(key) or "").strip():
                row[key] = default.get(key, "")
                changed += 1

        if row.get("active") is None:
            row["active"] = True
            changed += 1

        if row.get("critical_score") is None:
            row["critical_score"] = default.get("critical_score", 50)
            changed += 1

        notes = str(row.get("notes") or "").strip()
        if not notes or notes.upper().startswith("ENTER APPROVED"):
            row["notes"] = default.get("notes", "Starter rate - edit in Rate Master")
            changed += 1

    return rows, changed


def ensure_default_rates() -> tuple[list[dict], int]:
    current = load("rates", [])
    repaired, changed = _repair_default_rates(
        current if isinstance(current, list) else []
    )

    if changed or not isinstance(current, list):
        save("rates", repaired)

    return repaired, changed


def ensure_data():
    current_settings = load("settings", None)

    if not isinstance(current_settings, dict):
        save("settings", defaults_settings())
    else:
        merged_settings = {
            **defaults_settings(),
            **current_settings,
        }
        save("settings", merged_settings)

    # Always guarantee the complete built-in starter Rate Master.
    ensure_default_rates()

    for name, default in [
        ("extractions", []),
        ("reviews", []),
        ("quotations", []),
        ("revisions", []),
        (
            "dataset_meta",
            {
                "version": 1,
                "reviewed_at_version": 0,
            },
        ),
    ]:
        value = load(name, None)

        if value is None:
            save(name, default)


init_database(LEGACY_DATA_DIR)
ensure_data()

app = FastAPI(title="AI Manufacturing Quotation API", version="0.8.7")

_allowed_origins = [
    "http://localhost:3000",
    "http://127.0.0.1:3000",
]

_frontend_origin = os.getenv("FRONTEND_ORIGIN", "").strip().rstrip("/")
if _frontend_origin and _frontend_origin not in _allowed_origins:
    _allowed_origins.append(_frontend_origin)

app.add_middleware(
    CORSMiddleware,
    allow_origins=_allowed_origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


class Drawing(BaseModel):
    drawing_no: str
    revision: str
    description: str
    material: str
    thickness_mm: float
    weight_kg: float
    quantity: int
    notes: list[str] = Field(default_factory=list)


class RateItem(BaseModel):
    id: str = ""
    category: Literal["MATERIAL", "PROCESS", "LABOUR", "OTHER", "COMMERCIAL"]
    name: str
    grade: str = ""
    unit: str
    price: float = Field(ge=0)
    critical_score: int = Field(default=50, ge=0, le=100)
    active: bool = True
    notes: str = ""
    updated_at: str = ""


class Row(BaseModel):
    id: str
    category: str
    item: str
    drawingQty: str
    costingQty: float
    unit: str
    rate: float
    cost: float
    confidence: Literal["Exact", "Estimated", "Assumed"]
    rateId: str | None = None
    rateSource: str = "Manual / Included"
    criticalScore: int = Field(default=50, ge=0, le=100)


class CostReq(BaseModel):
    rows: list[Row]
    # Commercial percentages are controlled in Settings.
    # These legacy percentage fields remain accepted for backward compatibility,
    # but the UI no longer edits them directly on the costing sheet.
    material_wastage_pct: float | None = None
    overhead_pct: float | None = None
    markup_pct: float | None = None

    # Engineer-editable commercial amount overrides on the cost sheet.
    material_wastage_override: float | None = None
    overhead_override: float | None = None
    markup_override: float | None = None
    selling_price_override: float | None = None


class RateSyncReq(BaseModel):
    row: Row
    material_family: str = ""
    material_grade: str = ""
    material_specification: str = ""


class Summary(BaseModel):
    direct_cost: float
    material_wastage: float
    overhead: float
    manufacturing_cost: float
    markup: float
    selling_price: float
    material_wastage_pct: float = 0
    overhead_pct: float = 0
    markup_pct: float = 0
    material_wastage_critical: int = 0
    overhead_critical: int = 0
    markup_critical: int = 0


class ReviewReq(BaseModel):
    extraction_id: str
    file_hash: str
    drawing: Drawing
    rows: list[Row]


class RevisionReq(BaseModel):
    drawing: Drawing
    note: str = "Snapshot"


class QuoteReq(BaseModel):
    customer: str
    drawing: Drawing
    summary: Summary
    rows: list[Row] = Field(default_factory=list)
    status: str = "Draft"


class QuoteRenameReq(BaseModel):
    name: str = Field(min_length=1, max_length=160)


class ExportReq(BaseModel):
    drawing: Drawing
    rows: list[Row]
    summary: Summary


class PdfReq(ExportReq):
    customer: str = "Customer"


class BatchQuoteItem(BaseModel):
    drawing: Drawing
    rows: list[Row]
    summary: Summary


class BatchPdfReq(BaseModel):
    customer: str = "Customer"
    mode: Literal["separate", "merge"]
    items: list[BatchQuoteItem]


class BatchSaveReq(BaseModel):
    customer: str
    mode: Literal["separate", "merge"]
    items: list[BatchQuoteItem]
    status: str = "Draft"


class EngineeringArtifactReq(BaseModel):
    file_hash: str = ""
    filename: str = ""
    drawing: Drawing
    rows: list[Row] = Field(default_factory=list)
    ai_raw: dict = Field(default_factory=dict)


class Settings(BaseModel):
    company_name: str
    currency: str
    material_wastage_pct: float
    overhead_pct: float
    markup_pct: float
    auto_dataset_capture: bool
    learn_from_corrections: bool
    training_batch_threshold: int
    critical_medium_threshold: int = 40
    critical_high_threshold: int = 70


DFM_STANDARD_REFERENCES = [
    {
        "standard": "ISO 2768-1:1989",
        "scope": "General linear/angular tolerances where individual tolerances are not specified; applicable to metal-removal and formed sheet-metal parts. ISO 2768 edition 2 is in publication transition.",
    },
    {
        "standard": "ISO 1101:2017",
        "scope": "Geometrical tolerancing language for form, orientation, location and run-out.",
    },
    {
        "standard": "ISO 5458:2018",
        "scope": "Pattern and combined geometrical specifications; useful for repeated-hole/location patterns where applicable.",
    },
    {
        "standard": "ISO 21920-1:2021",
        "scope": "Surface-texture indication rules in technical product documentation.",
    },
    {
        "standard": "ISO 13715:2017",
        "scope": "Indication and dimensioning of edges of undefined shape / edge condition.",
    },
    {
        "standard": "ISO 2553:2019",
        "scope": "Symbolic representation of welded joints on technical drawings; current published edition while revision work is underway.",
    },
    {
        "standard": "ISO 9013:2017 + Amd 1:2024",
        "scope": "Thermal-cut classification, geometrical product specification and quality tolerances for oxyfuel/plasma/laser cuts where referenced.",
    },
    {
        "standard": "ISO 965-1:2026",
        "scope": "Tolerance system for ISO general-purpose metric screw threads (M).",
    },
]


def _artifact_process_names(ai_raw: dict, rows: list[Row]) -> list[str]:
    values: list[str] = []

    for item in ai_raw.get("manufacturing_processes") or []:
        if isinstance(item, dict):
            name = str(item.get("process") or item.get("name") or item.get("operation") or "").strip()
        else:
            name = str(item).strip()
        if name:
            values.append(name)

    for row in rows:
        if row.category.upper() == "PROCESS" and row.item:
            values.append(row.item)

    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        key = value.casefold()
        if key not in seen:
            seen.add(key)
            result.append(value)
    return result


def _dfm_classification(processes: list[str], ai_raw: dict) -> str:
    hay = " ".join(processes).casefold()
    fabrication_words = (
        "laser", "shear", "plasma", "waterjet", "bend", "forming",
        "weld", "grind", "polish", "passivation", "coat",
    )
    machining_words = (
        "turn", "mill", "machine", "drill", "bore", "tap",
        "thread", "chamfer", "ream", "cnc",
    )

    fabrication = any(word in hay for word in fabrication_words) or bool(ai_raw.get("bends") or ai_raw.get("welds"))
    machining = any(word in hay for word in machining_words) or bool(ai_raw.get("threads") or ai_raw.get("chamfers"))

    if fabrication and machining:
        return "Fabrication + Machining"
    if machining:
        return "Machining"
    if fabrication:
        return "Fabrication"
    return "Manufacturing route requires engineer review"


def _feature_number(item: object, keys: tuple[str, ...]) -> float | None:
    if not isinstance(item, dict):
        return None

    for key in keys:
        value = item.get(key)
        if value is None:
            continue
        match = re.search(r"-?\d+(?:\.\d+)?", str(value))
        if match:
            try:
                return float(match.group(0))
            except ValueError:
                pass
    return None


def _generate_dfm(payload: EngineeringArtifactReq) -> dict:
    drawing = payload.drawing
    ai_raw = payload.ai_raw or {}
    processes = _artifact_process_names(ai_raw, payload.rows)
    classification = _dfm_classification(processes, ai_raw)
    checks: list[dict] = []

    def add_check(area: str, result: str, finding: str, recommendation: str, standard: str = ""):
        checks.append({
            "area": area,
            "result": result,
            "finding": finding,
            "recommendation": recommendation,
            "standard": standard,
        })

    material_ok = bool(drawing.material and drawing.material.casefold() not in {"not detected", "unknown"})
    add_check(
        "Material definition",
        "PASS" if material_ok else "REVIEW",
        drawing.material if material_ok else "Material/grade is not fully defined.",
        "Confirm material family, grade and applicable material specification before release.",
    )

    if "Fabrication" in classification:
        add_check(
            "Sheet / plate thickness",
            "PASS" if drawing.thickness_mm > 0 else "FAIL",
            f"Thickness detected: {drawing.thickness_mm:g} mm." if drawing.thickness_mm > 0 else "Thickness is missing for a fabrication route.",
            "Verify thickness against selected stock and cutting/forming capability." if drawing.thickness_mm > 0 else "Add or confirm thickness before production planning.",
            "ISO 2768 / ISO 2768-1",
        )

    dimensions = ai_raw.get("dimensions") or []
    has_tolerance = any(
        isinstance(item, dict)
        and any(item.get(key) not in (None, "", "Not detected") for key in ("tolerance", "upper_tolerance", "lower_tolerance", "plus_minus"))
        for item in dimensions
    )
    add_check(
        "Dimensional tolerances",
        "PASS" if has_tolerance else "REVIEW",
        "Explicit tolerance information detected." if has_tolerance else "No explicit dimensional tolerance was detected in the extracted features.",
        "Confirm general tolerance note/class or add critical feature tolerances.",
        "ISO 2768 / ISO 2768-1",
    )

    confidence = ai_raw.get("confidence") or {}
    if isinstance(confidence, dict) and confidence:
        low_fields = [str(key) for key, value in confidence.items() if float(value or 0) < 70]
        add_check(
            "Drawing extraction confidence",
            "REVIEW" if low_fields else "PASS",
            f"Low-confidence fields: {', '.join(low_fields[:6])}" if low_fields else "No low-confidence extracted fields.",
            "Engineer-check every low-confidence feature before production release.",
        )

    if drawing.thickness_mm > 0 and ai_raw.get("holes"):
        risky_holes = []
        for item in ai_raw.get("holes") or []:
            diameter = _feature_number(item, ("diameter_mm", "diameter", "size_mm", "size"))
            if diameter is not None and diameter < 0.8 * drawing.thickness_mm:
                risky_holes.append(diameter)
        if risky_holes:
            add_check(
                "Small holes vs thickness",
                "REVIEW",
                f"{len(risky_holes)} hole(s) are smaller than ~0.8 × material thickness.",
                "Confirm cutting/drilling method, pierce quality and tolerance capability.",
            )

    if drawing.thickness_mm > 0 and ai_raw.get("bends"):
        tight_bends = []
        for item in ai_raw.get("bends") or []:
            radius = _feature_number(item, ("inside_radius_mm", "radius_mm", "radius", "r"))
            if radius is not None and radius < drawing.thickness_mm:
                tight_bends.append(radius)
        if tight_bends:
            add_check(
                "Bend radius",
                "REVIEW",
                f"{len(tight_bends)} bend-radius value(s) are below material thickness.",
                "Confirm material ductility, tooling radius and cracking risk.",
            )

    if ai_raw.get("welds"):
        weld_has_detail = any(
            isinstance(item, dict)
            and any(str(item.get(key) or "").strip() for key in ("type", "size", "length", "symbol", "process"))
            for item in ai_raw.get("welds") or []
        )
        add_check(
            "Weld definition / accessibility",
            "PASS" if weld_has_detail else "REVIEW",
            "Weld callouts detected." if weld_has_detail else "Weld features are present but the extracted weld definition is incomplete.",
            "Confirm weld type, size, access, sequence and inspection requirement.",
            "ISO 2553:2019",
        )

    if ai_raw.get("threads"):
        add_check(
            "Thread manufacturability",
            "REVIEW",
            f"{len(ai_raw.get('threads') or [])} threaded-feature callout(s) detected.",
            "Verify metric-thread tolerance class, engagement depth, tool access and edge distance.",
            "ISO 965-1:2026",
        )

    process_text = " ".join(processes).casefold()
    if any(name in process_text for name in ("laser", "plasma", "oxyfuel", "thermal cut")):
        add_check(
            "Thermal-cut quality",
            "REVIEW",
            "A thermal cutting process is included in the manufacturing route.",
            "Confirm cut-quality class, edge condition and dimensional tolerance required by the drawing/customer specification.",
            "ISO 9013:2017 + Amd 1:2024",
        )

    if len(ai_raw.get("holes") or []) >= 2:
        add_check(
            "Hole / feature pattern",
            "REVIEW",
            f"{len(ai_raw.get('holes') or [])} hole feature(s) detected; pattern/location capability should be confirmed.",
            "Verify positional scheme, datums and combined-pattern requirements where they are specified.",
            "ISO 5458:2018 / ISO 1101:2017",
        )

    surface_items = ai_raw.get("surface_finish") or []
    if surface_items:
        add_check(
            "Surface texture",
            "REVIEW",
            f"{len(surface_items)} surface-finish / texture callout(s) detected.",
            "Verify symbol interpretation, required parameter/value and whether the selected process can achieve it.",
            "ISO 21920-1:2021",
        )

    notes_text = " ".join(str(note) for note in (ai_raw.get("notes") or drawing.notes or []))
    if "sharp" in notes_text.casefold() or "deburr" in notes_text.casefold():
        add_check(
            "Edge condition",
            "PASS",
            "Edge/deburr requirement detected in drawing notes.",
            "Carry edge condition into the manufacturing and inspection plan.",
            "ISO 13715:2017",
        )
    else:
        add_check(
            "Edge condition",
            "REVIEW",
            "No explicit extracted edge-break/deburr requirement was found.",
            "Confirm whether sharp-edge removal / edge-break is required.",
            "ISO 13715:2017",
        )

    process_plan = []
    for index, process in enumerate(processes, 1):
        process_plan.append({
            "sequence": index,
            "process": process,
            "tooling": "Confirm machine/tooling from feature size, tolerance and material.",
            "feasibility": "PASS",
            "inspection": "Verify critical drawing features after this operation.",
        })

    fail_count = sum(1 for item in checks if item["result"] == "FAIL")
    review_count = sum(1 for item in checks if item["result"] == "REVIEW")
    status = "ATTENTION" if fail_count else ("REVIEW" if review_count else "READY")

    return {
        "id": f"DFM-{uuid.uuid4().hex[:10].upper()}",
        "created_at": now(),
        "name": f"DFM - {drawing.drawing_no or payload.filename or 'Drawing'}",
        "file_hash": payload.file_hash,
        "filename": payload.filename,
        "drawing_no": drawing.drawing_no,
        "revision": drawing.revision,
        "description": drawing.description,
        "classification": classification,
        "status": status,
        "standards": DFM_STANDARD_REFERENCES,
        "checks": checks,
        "process_plan": process_plan,
        "notes": [
            "Automated first-pass DFM screening.",
            "Final feasibility remains subject to engineer/tooling/machine capability review.",
        ],
    }


def _bom_item(item_no: int, category: str, description: str, material: str, specification: str, dimensions: str, quantity: float, unit: str, weight_kg: float, unit_cost: float, source: str) -> dict:
    return {
        "item_no": item_no,
        "category": category,
        "description": description,
        "material": material,
        "specification": specification,
        "dimensions": dimensions,
        "quantity": quantity,
        "unit": unit,
        "weight_kg": round(max(0.0, weight_kg), 4),
        "unit_cost": round(max(0.0, unit_cost), 2),
        "total_cost": round(max(0.0, quantity) * max(0.0, unit_cost), 2),
        "source": source,
        "remarks": "",
    }


def _generate_bom(payload: EngineeringArtifactReq) -> dict:
    drawing = payload.drawing
    ai_raw = payload.ai_raw or {}
    items: list[dict] = []

    material_row = next(
        (row for row in payload.rows if row.category.upper() == "MATERIAL" and row.id != "ai-material-stock"),
        None,
    )

    material_meta = ai_raw.get("material") or {}
    if not isinstance(material_meta, dict):
        material_meta = {}

    if material_row:
        items.append(_bom_item(
            1,
            "Raw Material",
            drawing.description or "Manufactured part material",
            drawing.material,
            str(material_meta.get("specification") or material_meta.get("grade") or ""),
            material_row.drawingQty,
            material_row.costingQty,
            material_row.unit,
            material_row.costingQty if material_row.unit.casefold() == "kg" else drawing.weight_kg,
            material_row.rate,
            "Drawing + Rate Master",
        ))

    item_no = len(items) + 1

    for stud in ai_raw.get("studs") or []:
        if not isinstance(stud, dict):
            continue
        qty = _feature_number(stud, ("quantity", "qty")) or 1
        size = str(stud.get("size") or stud.get("thread") or stud.get("diameter") or "")
        items.append(_bom_item(
            item_no,
            "Standard Part",
            "Stud / Fastener",
            str(stud.get("material") or ""),
            str(stud.get("standard") or ""),
            size,
            qty,
            "each",
            0,
            0,
            "Drawing feature",
        ))
        item_no += 1

    standard_part_pattern = re.compile(r"\b((?:M\d+(?:\.\d+)?)?\s*(?:bolt|nut|washer|screw|fastener))\b", re.IGNORECASE)
    found_standard: set[str] = set()

    for note in [str(note) for note in (ai_raw.get("notes") or drawing.notes or [])]:
        for match in standard_part_pattern.findall(note):
            clean = " ".join(match.split())
            key = clean.casefold()
            if key in found_standard:
                continue
            found_standard.add(key)
            items.append(_bom_item(
                item_no,
                "Standard Part",
                clean.title(),
                "",
                "",
                "",
                1,
                "each",
                0,
                0,
                "Drawing note",
            ))
            item_no += 1

    if not items:
        items.append(_bom_item(
            1,
            "Part",
            drawing.description or drawing.drawing_no or "Drawing item",
            drawing.material,
            "",
            "",
            max(1, drawing.quantity),
            "each",
            drawing.weight_kg,
            0,
            "Drawing",
        ))

    return {
        "id": f"BOM-{uuid.uuid4().hex[:10].upper()}",
        "created_at": now(),
        "name": f"BOM - {drawing.drawing_no or payload.filename or 'Drawing'}",
        "file_hash": payload.file_hash,
        "filename": payload.filename,
        "drawing_no": drawing.drawing_no,
        "revision": drawing.revision,
        "description": drawing.description,
        "status": "READY",
        "items": items,
        "notes": [
            "Generated from drawing extraction and current costing rows.",
            "Engineer-verify standard-part quantity/cost where the drawing is not explicit.",
        ],
    }


def rates_list() -> list[dict]:
    rows = load("rates", default_rate_items())
    return rows if isinstance(rows, list) else default_rate_items()


def rate_by_id(rate_id: str | None) -> dict | None:
    if not rate_id:
        return None
    return next((r for r in rates_list() if r.get("id") == rate_id and r.get("active", True)), None)


def find_rate(category: str, name: str, grade: str = "") -> dict | None:
    candidates = [
        r
        for r in rates_list()
        if r.get("active", True)
        and r.get("category") == category
        and r.get("name", "").casefold() == name.casefold()
    ]
    if grade:
        exact = next(
            (r for r in candidates if r.get("grade", "").casefold() == grade.casefold()),
            None,
        )
        if exact:
            return exact
    return candidates[0] if candidates else None


def row_from_rate(
    row_id: str,
    category: str,
    item: str,
    drawing_qty: str,
    costing_qty: float,
    confidence: str,
    rate_id: str | None,
    fallback_unit: str,
    fallback_rate: float = 0,
    fallback_critical: int = 30,
    fallback_source: str = "Included / Manual",
) -> Row:
    saved = rate_by_id(rate_id)
    if saved:
        rate = float(saved.get("price", 0))
        unit = saved.get("unit", fallback_unit)
        critical = int(saved.get("critical_score", 50))
        source = "Rate Master"
    else:
        rate = fallback_rate
        unit = fallback_unit
        critical = fallback_critical
        source = fallback_source
    return Row(
        id=row_id,
        category=category,
        item=item,
        drawingQty=drawing_qty,
        costingQty=costing_qty,
        unit=unit,
        rate=rate,
        cost=costing_qty * rate,
        confidence=confidence,
        rateId=rate_id if saved else None,
        rateSource=source,
        criticalScore=critical,
    )


def calc(
    rows: list[Row],
    material_wastage_pct: float | None = None,
    overhead_pct: float | None = None,
    markup_pct: float | None = None,
    material_wastage_override: float | None = None,
    overhead_override: float | None = None,
    markup_override: float | None = None,
    selling_price_override: float | None = None,
) -> Summary:
    """
    Costing rules:
    - Material/process/labour values come from the cost rows / Rate Master.
    - Commercial PERCENTAGES come from Settings.
    - Commercial AMOUNTS may be manually overridden in the cost sheet.
    - Legacy explicit percentage inputs are accepted only for compatibility.
    """
    settings = load("settings", defaults_settings())

    waste_rate = find_rate("COMMERCIAL", "Material Wastage")
    overhead_rate = find_rate("COMMERCIAL", "Factory Overhead")
    markup_rate = find_rate("COMMERCIAL", "Profit / Markup")

    # Settings are the user-controlled source of truth for these percentages.
    default_waste = float(settings["material_wastage_pct"])
    default_overhead = float(settings["overhead_pct"])
    default_markup = float(settings["markup_pct"])

    # Backward-compatible API callers can still explicitly provide a percentage.
    waste_pct = (
        max(0.0, float(material_wastage_pct))
        if material_wastage_pct is not None
        else default_waste
    )
    overhead_pct_value = (
        max(0.0, float(overhead_pct))
        if overhead_pct is not None
        else default_overhead
    )
    markup_pct_value = (
        max(0.0, float(markup_pct))
        if markup_pct is not None
        else default_markup
    )

    direct = sum(
        max(0, r.costingQty) * max(0, r.rate)
        for r in rows
    )

    material_for_wastage = sum(
        max(0, r.costingQty) * max(0, r.rate)
        for r in rows
        if r.category.upper() == "MATERIAL"
        and r.id != "ai-material-stock"
    )

    automatic_wastage = material_for_wastage * waste_pct / 100
    wastage = (
        max(0.0, float(material_wastage_override))
        if material_wastage_override is not None
        else automatic_wastage
    )

    automatic_overhead = (direct + wastage) * overhead_pct_value / 100
    overhead = (
        max(0.0, float(overhead_override))
        if overhead_override is not None
        else automatic_overhead
    )

    manufacturing = direct + wastage + overhead

    automatic_markup = manufacturing * markup_pct_value / 100
    markup = (
        max(0.0, float(markup_override))
        if markup_override is not None
        else automatic_markup
    )

    selling_price = manufacturing + markup

    if selling_price_override is not None:
        selling_price = max(0.0, float(selling_price_override))

    return Summary(
        direct_cost=round(direct, 2),
        material_wastage=round(wastage, 2),
        overhead=round(overhead, 2),
        manufacturing_cost=round(manufacturing, 2),
        markup=round(markup, 2),
        selling_price=round(selling_price, 2),
        material_wastage_pct=round(waste_pct, 2),
        overhead_pct=round(overhead_pct_value, 2),
        markup_pct=round(markup_pct_value, 2),
        material_wastage_critical=int(
            waste_rate.get("critical_score", 50)
        ) if waste_rate else 50,
        overhead_critical=int(
            overhead_rate.get("critical_score", 50)
        ) if overhead_rate else 50,
        markup_critical=int(
            markup_rate.get("critical_score", 50)
        ) if markup_rate else 50,
    )


def clean_filename_description(filename: str) -> str:
    stem = Path(filename or "drawing").stem
    stem = re.sub(r"^\[?PID[#\s_-]*\d+\]?\s*", "", stem, flags=re.I)
    stem = re.sub(r"\s*-\s*R\d+(?:\.\d+)?(?:-\d{4}-\d{2}-\d{2})?\s*$", "", stem, flags=re.I)
    return re.sub(r"[_]+", " ", stem).strip() or "Engineering Drawing"


def extract_pdf_text(content: bytes) -> str:
    try:
        reader = PdfReader(BytesIO(content))
        chunks = []
        for page in reader.pages:
            try:
                chunks.append(page.extract_text() or "")
            except Exception:
                continue
        return "\n".join(chunks)
    except Exception:
        return ""


def first_match(patterns: list[str], text: str, flags=re.I | re.M):
    for pattern in patterns:
        match = re.search(pattern, text, flags)
        if match:
            return match
    return None


def detect_material(text: str, filename: str):
    hay = f"{filename}\n{text}"

    ss = first_match(
        [
            r"\bAISI\s*(201|202|304L?|316L?|321|430)\b",
            r"\b(201|202|304L?|316L?|321|430)\s*SS\b",
            r"\bSS\s*(201|202|304L?|316L?|321|430)\b",
        ],
        hay,
    )
    if ss:
        grade = ss.group(1).upper()
        rate = find_rate("MATERIAL", "Stainless Steel", f"AISI {grade}")
        return f"AISI {grade} Stainless Steel", (rate or {}).get("id")

    ms = first_match([r"\b(?:IS\s*2062\s*)?(E250|E350)\b"], hay)
    if ms:
        grade = ms.group(1).upper()
        rate = find_rate("MATERIAL", "Mild Steel", f"IS 2062 {grade}")
        return f"Mild Steel IS 2062 {grade}", (rate or {}).get("id")

    if re.search(r"\b(?:MILD\s*STEEL|\bMS\b)", hay, re.I):
        return "Mild Steel", None

    al = first_match([r"\b(5052(?:-H32)?|6061(?:-T6)?|6082(?:-T6)?)\b"], hay)
    if al and re.search(r"\bAL(?:UMINIUM|UMINUM)?\b", hay, re.I):
        grade = al.group(1).upper()
        family = "Aluminium"
        grade_lookup = {
            "5052": "5052-H32", "5052-H32": "5052-H32",
            "6061": "6061-T6", "6061-T6": "6061-T6",
            "6082": "6082-T6", "6082-T6": "6082-T6",
        }.get(grade, grade)
        rate = find_rate("MATERIAL", family, grade_lookup)
        return f"{family} {grade_lookup}", (rate or {}).get("id")

    return "Not detected", None


def parse_float_match(match, group=1, default=0.0) -> float:
    if not match:
        return default
    try:
        return float(match.group(group).replace(",", "."))
    except Exception:
        return default


def extract_drawing_fields(content: bytes, filename: str):
    suffix = Path(filename or "").suffix.lower()
    warnings: list[str] = []
    raw_text = ""

    if suffix == ".pdf":
        raw_text = extract_pdf_text(content)
        source = "pdf_parser"
        if len(raw_text.strip()) < 20:
            warnings.append(
                "PDF text could not be read reliably. AI/Vision OCR is not connected yet; review all fields manually."
            )
    else:
        source = "file_specific_blank"
        warnings.append(
            f"{suffix or 'This file type'} needs the AI/CAD extraction module. "
            "The app intentionally does not reuse another drawing's data."
        )

    hay = f"{filename}\n{raw_text}"
    drawing_match = first_match(
        [
            r"\bPID\s*[#:\-]?\s*(\d{3,})\b",
            r"\[PID#(\d+)\]",
            r"\bDRAWING\s*(?:NO\.?|NUMBER)?\s*[:#\-]?\s*([A-Z0-9._/-]{3,})",
        ],
        hay,
    )
    drawing_no = drawing_match.group(1).strip() if drawing_match else ""
    if not drawing_no:
        file_pid = re.search(r"PID[#\s_-]*(\d+)", filename or "", re.I)
        drawing_no = file_pid.group(1) if file_pid else Path(filename or "drawing").stem[:60]

    rev_match = first_match(
        [
            r"\bREV(?:ISION)?\s*[:#\-]?\s*(R?\d+(?:\.\d+)?)\b",
            r"\b(R\d+(?:\.\d+)?)\b",
        ],
        hay,
    )
    revision = rev_match.group(1).upper() if rev_match else ""

    material, material_rate_id = detect_material(raw_text, filename)

    thickness_match = first_match(
        [
            r"\bTHICKNESS\s*[:=]?\s*(\d+(?:[.,]\d+)?)\s*(?:MM)?\b",
            r"\b(\d+(?:[.,]\d+)?)\s*MM\s*(?:THK|THICK)\b",
            r"\b(\d+(?:[.,]\d+)?)\s*(?:THK|THICK)\b",
        ],
        hay,
    )
    thickness = parse_float_match(thickness_match)

    weight_match = first_match(
        [
            r"\bWEIGHT\s*[:=]?\s*(\d+(?:[.,]\d+)?)\s*KG\b",
            r"\b(\d+(?:[.,]\d+)?)\s*KG\s*(?:WEIGHT)?\b",
        ],
        hay,
    )
    weight = parse_float_match(weight_match)

    qty_match = first_match(
        [
            r"\bQTY(?:\.|UANTITY)?\s*[:=]?\s*(\d+)\b",
            r"\bQUANTITY\s*[:=]?\s*(\d+)\b",
        ],
        hay,
    )
    qty = int(qty_match.group(1)) if qty_match else 1

    description = clean_filename_description(filename)
    desc_match = first_match(
        [
            r"\bDESCRIPTION\s*[:=]\s*([^\n\r]{3,100})",
            r"\bTITLE\s*[:=]\s*([^\n\r]{3,100})",
        ],
        raw_text,
    )
    if desc_match:
        description = desc_match.group(1).strip()

    notes = []
    note_rules = [
        (r"FULLY\s*WELD(?:ED|ING)", "Fully welded fabrication"),
        (r"GRIND(?:ING)?\s*(?:AND|&)?\s*FLUSH", "Grind and flush welded edges"),
        (r"DEBURR", "Deburr all sharp edges"),
        (r"\bTIG\b", "TIG welding specified"),
        (r"\bMIG\b", "MIG welding specified"),
        (r"\bPASSIVAT", "Passivation specified"),
        (r"\bPOLISH", "Polishing specified"),
        (r"\bPOWDER\s*COAT", "Powder coating specified"),
    ]
    for pattern, label in note_rules:
        if re.search(pattern, hay, re.I):
            notes.append(label)

    hole_match = first_match(
        [
            r"(?:Ø|DIA(?:METER)?\.?)\s*(\d+(?:[.,]\d+)?)\s*(?:MM)?[^\n\r]{0,40}?\b(\d+)\s*(?:PLS|PLACES|NOS?)\b",
            r"\b(\d+)\s*(?:PLS|PLACES)\b[^\n\r]{0,40}?(?:Ø|DIA)",
        ],
        hay,
    )
    hole_count = 0
    hole_dia = 0.0
    if hole_match:
        try:
            if hole_match.lastindex and hole_match.lastindex >= 2:
                hole_dia = float(hole_match.group(1).replace(",", "."))
                hole_count = int(hole_match.group(2))
            else:
                hole_count = int(hole_match.group(1))
        except Exception:
            pass
    if hole_count:
        notes.append(
            f"Detected hole pattern: {hole_count} place(s)"
            + (f", Ø{hole_dia:g} mm" if hole_dia else "")
        )

    stud_match = first_match(
        [
            r"\b(\d+)\s*[X×]\s*[^\n\r]{0,30}?\bM(\d+)\s*[X×]\s*(\d+(?:[.,]\d+)?)\s*MM\b",
            r"\b(\d+)\s*(?:STUDS?|PLS)[^\n\r]{0,30}?\bM(\d+)\b",
        ],
        hay,
    )
    stud_count = int(stud_match.group(1)) if stud_match else 0
    if stud_count:
        notes.append(f"Detected studs: {stud_count}")

    drawing = Drawing(
        drawing_no=drawing_no,
        revision=revision,
        description=description,
        material=material,
        thickness_mm=thickness,
        weight_kg=weight,
        quantity=max(1, qty),
        notes=notes,
    )

    rows: list[Row] = []

    if material != "Not detected":
        material_item = material.replace(" Stainless Steel", " SS")
        if weight > 0:
            rows.append(
                row_from_rate(
                    "mat",
                    "MATERIAL",
                    material_item,
                    f"{weight:g} kg net",
                    weight,
                    "Exact",
                    material_rate_id,
                    "kg",
                    fallback_critical=95,
                    fallback_source="Material detected; rate not found",
                )
            )
        else:
            rows.append(
                row_from_rate(
                    "mat",
                    "MATERIAL",
                    material_item,
                    "Weight not detected",
                    0,
                    "Estimated",
                    material_rate_id,
                    "kg",
                    fallback_critical=100,
                    fallback_source="Weight needs review",
                )
            )

    upper = hay.upper()

    if "LASER" in upper:
        rows.append(row_from_rate("laser", "PROCESS", "Laser Cutting", "Detected", 1, "Estimated", "PROC-LASER", "job"))

    if re.search(r"\bBEND(?:ING)?\b|\bPRESS\s*BRAKE\b|\bFORM(?:ING)?\b", upper):
        rows.append(row_from_rate("bend", "PROCESS", "Press Brake Forming", "Detected", 1, "Estimated", "PROC-BEND", "job"))

    if hole_count:
        drill_rate = rate_by_id("PROC-DRILL")
        rows.append(
            row_from_rate(
                "holes",
                "PROCESS",
                f"Drilling / holes" + (f" Ø{hole_dia:g} mm" if hole_dia else ""),
                f"{hole_count} holes",
                hole_count,
                "Exact",
                "PROC-DRILL" if drill_rate else None,
                "hole",
                0,
                60,
                "Hole count detected; verify manufacturing method",
            )
        )

    if "TIG" in upper:
        rows.append(row_from_rate("weld", "PROCESS", "TIG Welding", "Weld length not detected", 0, "Estimated", "PROC-TIG", "m", fallback_critical=100))
    elif "MIG" in upper:
        rows.append(row_from_rate("weld", "PROCESS", "MIG Welding", "Weld length not detected", 0, "Estimated", "PROC-MIG", "m", fallback_critical=100))
    elif "WELD" in upper:
        rows.append(
            row_from_rate(
                "weld",
                "PROCESS",
                "Welding — type/length review required",
                "Welding detected",
                0,
                "Estimated",
                None,
                "m",
                0,
                100,
                "Needs weld type + length",
            )
        )

    if stud_count:
        rows.append(row_from_rate("stud", "PROCESS", "Stud Welding", f"{stud_count} studs", stud_count, "Exact", "PROC-STUD", "stud"))

    if re.search(r"GRIND", upper):
        rows.append(row_from_rate("grind", "PROCESS", "Grinding & Flush", "Detected", 1, "Exact", "PROC-GRIND", "job"))

    if re.search(r"DEBURR", upper):
        rows.append(row_from_rate("deburr", "PROCESS", "Deburring", "Detected", 1, "Exact", "PROC-DEBURR", "job"))

    if re.search(r"POLISH", upper):
        rows.append(row_from_rate("polish", "PROCESS", "Polishing", "Area not detected", 0, "Estimated", "PROC-POLISH", "m2"))

    if re.search(r"PASSIVAT", upper):
        rows.append(row_from_rate("passivation", "PROCESS", "Passivation", "Detected", 1, "Estimated", "PROC-PASS", "job"))

    if re.search(r"POWDER\s*COAT", upper):
        rows.append(row_from_rate("coat", "PROCESS", "Powder Coating", "Area not detected", 0, "Estimated", "PROC-COAT", "m2"))

    # Labour is deliberately conservative: only add when a related process is detected.
    if any(r.category == "PROCESS" for r in rows):
        rows.append(row_from_rate("fab", "LABOUR", "Fabricator", "Estimated", 1, "Assumed", "LAB-FAB", "hr"))

    if any("TIG" in r.item for r in rows):
        rows.append(row_from_rate("tig-labour", "LABOUR", "TIG Welder", "Estimated", 1, "Assumed", "LAB-TIG", "hr"))
    elif any("MIG" in r.item for r in rows):
        rows.append(row_from_rate("mig-labour", "LABOUR", "MIG Welder", "Estimated", 1, "Assumed", "LAB-MIG", "hr"))

    if any(("Grinding" in r.item or "Deburring" in r.item or "Polishing" in r.item) for r in rows):
        rows.append(row_from_rate("finish-labour", "LABOUR", "Finishing Operator", "Estimated", 0.5, "Assumed", "LAB-FINISH", "hr"))

    if not raw_text.strip() and suffix == ".pdf":
        source = "pdf_text_unreadable"

    if not rows:
        warnings.append(
            "No reliable costing rows were detected. This is intentional: the system will not copy a previous drawing's values."
        )

    return drawing, rows, source, warnings, raw_text[:1500]


def ai_result_to_drawing(ai_data: dict) -> Drawing:
    """Convert Vision-AI JSON into the Drawing model used by the web app."""

    material_data = ai_data.get("material") or {}
    family = str(material_data.get("family") or "").strip()
    grade = str(material_data.get("grade") or "").strip()
    specification = str(material_data.get("specification") or "").strip()

    material_parts = []
    for value in [family, grade, specification]:
        if value and value.casefold() not in {x.casefold() for x in material_parts}:
            material_parts.append(value)

    material = " ".join(material_parts).strip() or "Not detected"

    notes: list[str] = []

    for note in ai_data.get("notes") or []:
        if note:
            notes.append(str(note).strip())

    for process in ai_data.get("manufacturing_processes") or []:
        if isinstance(process, dict):
            process_name = str(process.get("process") or "").strip()
            confidence = process.get("confidence")
            if process_name:
                suffix = f" ({confidence}% confidence)" if confidence is not None else ""
                notes.append(f"Process: {process_name}{suffix}")

    for item in ai_data.get("missing_or_uncertain") or []:
        if item:
            notes.append(f"Review required: {item}")

    # Preserve important engineering features in the current Review UI.
    # The frontend model does not yet have dedicated tables for these fields,
    # so they are shown as structured notes instead of being discarded.
    dimensions = ai_data.get("dimensions") or []
    for dim in dimensions:
        if not isinstance(dim, dict):
            continue
        value = dim.get("value_mm")
        label = str(dim.get("label") or dim.get("type") or "").strip()
        tolerance = str(dim.get("tolerance") or "").strip()
        qty = int(dim.get("quantity") or 1)
        if value is not None:
            detail = f"Dimension: {label + ' ' if label else ''}{value} mm"
            if tolerance:
                detail += f" ({tolerance})"
            if qty > 1:
                detail += f" × {qty}"
            notes.append(detail)

    threads = ai_data.get("threads") or []
    for thread in threads:
        if not isinstance(thread, dict):
            continue
        designation = str(thread.get("designation") or "").strip()
        qty = int(thread.get("quantity") or 0)
        through = bool(thread.get("through"))
        if designation:
            detail = f"Thread: {designation}"
            if qty:
                detail += f" × {qty}"
            if through:
                detail += " THRU"
            notes.append(detail)

    chamfers = ai_data.get("chamfers") or []
    for chamfer in chamfers:
        if not isinstance(chamfer, dict):
            continue
        size = chamfer.get("size_mm")
        angle = chamfer.get("angle_deg")
        qty = int(chamfer.get("quantity") or 0)
        if size is not None or angle is not None:
            detail = "Chamfer:"
            if size is not None:
                detail += f" {size} mm"
            if angle is not None:
                detail += f" × {angle}°"
            if qty:
                detail += f" ({qty} place(s))"
            notes.append(detail)

    bends = ai_data.get("bends") or []
    for bend in bends:
        if not isinstance(bend, dict):
            continue
        angle = bend.get("angle_deg")
        qty = int(bend.get("quantity") or 0)
        if angle is not None:
            notes.append(f"Bend: {angle}° × {qty or 1}")

    for finish in ai_data.get("surface_finish") or []:
        if finish:
            notes.append(f"Surface finish: {finish}")

    holes = ai_data.get("holes") or []
    for hole in holes:
        if not isinstance(hole, dict):
            continue
        qty = int(hole.get("quantity") or 0)
        dia = hole.get("diameter_mm")
        hole_type = str(hole.get("type") or "").strip()
        if qty:
            detail = f"Holes: {qty}"
            if dia is not None:
                detail += f" × Ø{dia} mm"
            if hole_type:
                detail += f" {hole_type}"
            notes.append(detail)

    studs = ai_data.get("studs") or []
    for stud in studs:
        if not isinstance(stud, dict):
            continue
        qty = int(stud.get("quantity") or 0)
        size = str(stud.get("size") or "").strip()
        length = stud.get("length_mm")
        if qty:
            detail = f"Studs: {qty}"
            if size:
                detail += f" × {size}"
            if length is not None:
                detail += f" × {length} mm"
            notes.append(detail)

    welds = ai_data.get("welds") or []
    for weld in welds:
        if not isinstance(weld, dict):
            continue
        weld_type = str(weld.get("type") or "").strip()
        length = weld.get("length_mm")
        location = str(weld.get("location") or "").strip()
        if weld_type or length is not None or location:
            detail = "Weld"
            if weld_type:
                detail += f": {weld_type}"
            if length is not None:
                detail += f", {length} mm"
            if location:
                detail += f", {location}"
            notes.append(detail)

    def safe_float(value, default=0.0):
        try:
            return float(value) if value is not None else default
        except (TypeError, ValueError):
            return default

    def safe_int(value, default=1):
        try:
            parsed = int(value) if value is not None else default
            return max(1, parsed)
        except (TypeError, ValueError):
            return default

    return Drawing(
        drawing_no=str(ai_data.get("drawing_no") or "").strip(),
        revision=str(ai_data.get("revision") or "").strip(),
        description=str(ai_data.get("description") or "Engineering Drawing").strip(),
        material=material,
        thickness_mm=safe_float(ai_data.get("thickness_mm")),
        weight_kg=safe_float(ai_data.get("weight_kg")),
        quantity=safe_int(ai_data.get("product_quantity"), 1),
        notes=notes,
    )


def _norm(value: str) -> str:
    return re.sub(r"[^A-Z0-9]+", "", str(value or "").upper())


def find_material_rate_from_ai(ai_data: dict) -> dict | None:
    """Match AI material family/grade/specification to an active MATERIAL rate."""
    material = ai_data.get("material") or {}
    family = str(material.get("family") or "")
    grade = str(material.get("grade") or "")
    spec = str(material.get("specification") or "")

    combined = _norm(f"{family} {grade} {spec}")

    # Strong known-grade aliases first.
    aliases = {
        "S355J0": "MAT-CS-S355J0",
        "10553": "MAT-CS-S355J0",
        "AISI304": "MAT-SS-304",
        "304SS": "MAT-SS-304",
        "AISI304L": "MAT-SS-304L",
        "AISI316": "MAT-SS-316",
        "AISI316L": "MAT-SS-316L",
        "AISI430": "MAT-SS-430",
        "IS2062E250": "MAT-MS-E250",
        "IS2062E350": "MAT-MS-E350",
        "ASTMA36": "MAT-CS-A36",
        "6061T6": "MAT-AL-6061",
        "6082T6": "MAT-AL-6082",
        "5052H32": "MAT-AL-5052",
    }

    for alias, rate_id in aliases.items():
        if alias in combined:
            found = rate_by_id(rate_id)
            if found:
                return found

    # Generic grade/name comparison.
    for rate in rates_list():
        if rate.get("category") != "MATERIAL" or not rate.get("active", True):
            continue
        rate_grade = _norm(rate.get("grade", ""))
        rate_name = _norm(rate.get("name", ""))
        if rate_grade and (rate_grade in combined or combined in rate_grade):
            return rate
        if rate_name and rate_name in combined and not grade:
            return rate

    return None


def ai_confidence_label(score) -> Literal["Exact", "Estimated", "Assumed"]:
    try:
        value = float(score)
    except (TypeError, ValueError):
        value = 50.0

    if value >= 85:
        return "Exact"
    if value >= 60:
        return "Estimated"
    return "Assumed"


def process_rate_id(process_name: str) -> str | None:
    """Map AI process wording to Rate Master IDs, including user-added processes."""
    name = _norm(process_name)

    # User-maintained Rate Master is the source of truth. If Vision AI returns
    # the exact process name saved by the engineer, use that rate directly.
    for rate in rates_list():
        if (
            rate.get("active", True)
            and rate.get("category") == "PROCESS"
            and _norm(str(rate.get("name") or "")) == name
        ):
            return str(rate.get("id") or "") or None

    rules = [
        (("LASER",), "PROC-LASER"),
        (("SHEAR", "GUILLOTINE"), "PROC-SHEAR"),
        (("PLASMA",), "PROC-PLASMA"),
        (("WATERJET", "WATERJETCUT"), "PROC-WATERJET"),
        (("PRESSBRAKE", "BEND", "FORMING"), "PROC-BEND"),
        (("TIG",), "PROC-TIG"),
        (("MIG",), "PROC-MIG"),
        (("STUDWELD",), "PROC-STUD"),
        (("TACKWELD", "TACKWELDING", "GENERALWELD", "WELDING", "WELD"), "PROC-WELD"),
        (("GRIND",), "PROC-GRIND"),
        (("DEBURR",), "PROC-DEBURR"),
        (("PASSIV",), "PROC-PASS"),
        (("POLISH",), "PROC-POLISH"),
        (("POWDERCOAT",), "PROC-COAT"),
        (("INSPECT", "QC"), "PROC-QC"),
        (("SAW", "RAWCUT", "STOCKCUT"), "PROC-SAW"),
        (("CNCTURN", "TURNING", "LATHE"), "PROC-TURN"),
        (("CNCMILL", "MILLING"), "PROC-MILL"),
        (("MANUALMILL",), "PROC-MANUAL-MILL"),
        (("HANDGRIND", "CUTOFF", "CUT-OFF"), "PROC-HAND-GRIND"),
        (("BORING", "BORE"), "PROC-BORE"),
        (("DRILL",), "PROC-BORE"),
        (("THREAD", "TAPPING", "TAP"), "PROC-THREAD"),
        (("CHAMFER",), "PROC-CHAMFER"),
        (("MACHINING", "MACHINE"), "PROC-MACHINE"),
    ]

    for keys, rate_id in rules:
        if any(key in name for key in keys):
            return rate_id
    return None



# Internal stock assumptions used only for costing.
# They are not displayed in the application UI.
#
# Where we have a useful flat-sheet stock catalog, the engine chooses the
# smallest economical purchasable format. For other material families it uses
# the exact required blank rather than inventing an unknown supplier size.
_INTERNAL_STOCK_SIZES_MM = {
    "STAINLESS": [
        (500.0, 500.0),
        (1000.0, 2000.0),
        (1250.0, 2500.0),
        (1500.0, 3000.0),
        (2000.0, 4000.0),
    ],
    "ALUMINIUM": [
        (500.0, 500.0),
        (1000.0, 2000.0),
        (1250.0, 2500.0),
        (1500.0, 3000.0),
        (1525.0, 3660.0),
        (2000.0, 4000.0),
        (2000.0, 6000.0),
    ],
    "ALUMINUM": [
        (500.0, 500.0),
        (1000.0, 2000.0),
        (1250.0, 2500.0),
        (1500.0, 3000.0),
        (1525.0, 3660.0),
        (2000.0, 4000.0),
        (2000.0, 6000.0),
    ],
    "COPPER": [
        (500.0, 500.0),
        (600.0, 1220.0),
        (1000.0, 2000.0),
        (1220.0, 2440.0),
        (1250.0, 2500.0),
        (1500.0, 3000.0),
        (2000.0, 4000.0),
    ],
    "MILDSTEEL": [
        (500.0, 500.0),
        (1000.0, 2000.0),
        (1250.0, 2500.0),
        (1500.0, 3000.0),
    ],
    "CARBONSTEEL": [
        (500.0, 500.0),
        (1000.0, 2000.0),
        (1250.0, 2500.0),
        (1500.0, 3000.0),
    ],
    "STRUCTURALSTEEL": [
        (500.0, 500.0),
        (1000.0, 2000.0),
        (1250.0, 2500.0),
        (1500.0, 3000.0),
    ],
    "GALVANIZEDSTEEL": [
        (500.0, 500.0),
        (1000.0, 2000.0),
        (1250.0, 2500.0),
        (1500.0, 3000.0),
    ],
    "BRASS": [
        (500.0, 500.0),
        (1000.0, 2000.0),
        (1220.0, 2440.0),
    ],
}

_INTERNAL_DENSITY_KG_PER_MM3 = {
    "STAINLESS": 7.93e-6,
    "MILDSTEEL": 7.85e-6,
    "CARBONSTEEL": 7.85e-6,
    "STRUCTURALSTEEL": 7.85e-6,
    "GALVANIZEDSTEEL": 7.85e-6,
    "ALUMINIUM": 2.70e-6,
    "ALUMINUM": 2.70e-6,
    "COPPER": 8.96e-6,
    "BRASS": 8.50e-6,
}


def _internal_material_key(ai_data: dict, drawing: Drawing) -> str:
    material = ai_data.get("material") or {}
    return _norm(
        f"{material.get('family', '')} "
        f"{material.get('grade', '')} "
        f"{material.get('specification', '')} "
        f"{drawing.material}"
    )


def _internal_density(ai_data: dict, drawing: Drawing) -> float:
    key = _internal_material_key(ai_data, drawing)
    for family, density in _INTERNAL_DENSITY_KG_PER_MM3.items():
        if family in key:
            return density
    return 7.85e-6


def _internal_stock_catalog(ai_data: dict, drawing: Drawing) -> list[tuple[float, float]]:
    key = _internal_material_key(ai_data, drawing)
    for family, sizes in _INTERNAL_STOCK_SIZES_MM.items():
        if family in key:
            return sizes
    return []


def _is_sheet_based_part(ai_data: dict, drawing: Drawing) -> bool:
    try:
        thickness = float(ai_data.get("thickness_mm") or drawing.thickness_mm or 0)
    except (TypeError, ValueError):
        thickness = 0

    if thickness <= 0:
        return False

    processes = " ".join(
        str(x.get("process") or "")
        for x in (ai_data.get("manufacturing_processes") or [])
        if isinstance(x, dict)
    )

    combined = _norm(
        f"{drawing.description} {drawing.material} "
        f"{processes} {' '.join(str(x) for x in (ai_data.get('notes') or []))}"
    )

    if any(
        signal in combined
        for signal in [
            "SHEET", "PLATE", "LASER", "PRESSBRAKE", "BEND",
            "FORMING", "STRUT", "TRAY", "PANEL", "BRACKET", "COVER"
        ]
    ):
        return True

    usable_dims = []
    for dim in ai_data.get("dimensions") or []:
        if not isinstance(dim, dict):
            continue
        try:
            value = float(dim.get("value_mm") or 0)
        except (TypeError, ValueError):
            continue
        if value > max(20.0, thickness * 3):
            usable_dims.append(value)

    return len(usable_dims) >= 2


def _internal_part_length_width(
    ai_data: dict,
    drawing: Drawing,
) -> tuple[float, float] | None:
    try:
        thickness = float(ai_data.get("thickness_mm") or drawing.thickness_mm or 0)
    except (TypeError, ValueError):
        thickness = 0

    dimension_rows: list[tuple[float, str]] = []

    for dim in ai_data.get("dimensions") or []:
        if not isinstance(dim, dict):
            continue

        try:
            value = float(dim.get("value_mm") or 0)
        except (TypeError, ValueError):
            continue

        if value <= max(20.0, thickness * 3):
            continue

        label = str(dim.get("label") or dim.get("type") or "").strip()
        dimension_rows.append((value, label))

    if not dimension_rows:
        return None

    preferred = [
        (value, label)
        for value, label in dimension_rows
        if any(
            key in _norm(label)
            for key in ["OVERALL", "LENGTH", "WIDTH", "HEIGHT", "OAL"]
        )
    ]

    source_rows = preferred if len(preferred) >= 2 else dimension_rows
    values = sorted(
        {round(value, 3) for value, _ in source_rows if value > 0},
        reverse=True,
    )

    if not values:
        return None

    length = values[0]
    width = values[1] if len(values) > 1 else values[0]
    return max(1.0, length), max(1.0, width)


_MATERIAL_BLANK_MARGIN_MM = 100.0


def _internal_sheet_material_basis(
    ai_data: dict,
    drawing: Drawing,
) -> tuple[float, float, float] | None:
    """
    Cost sheet/plate material from the actual extracted part envelope plus a
    simple 100 mm allowance on BOTH overall dimensions.

    Example:
      drawing envelope = 300 x 400 mm
      costing blank    = 400 x 500 mm

    We intentionally do NOT charge a full supplier sheet here. The user can
    still edit the resulting kg quantity/rate in the engineering cost sheet.
    """
    if not _is_sheet_based_part(ai_data, drawing):
        return None

    try:
        thickness = float(ai_data.get("thickness_mm") or drawing.thickness_mm or 0)
    except (TypeError, ValueError):
        return None

    if thickness <= 0:
        return None

    size = _internal_part_length_width(ai_data, drawing)
    if not size:
        return None

    part_length, part_width = size
    blank_length = max(1.0, part_length + _MATERIAL_BLANK_MARGIN_MM)
    blank_width = max(1.0, part_width + _MATERIAL_BLANK_MARGIN_MM)

    product_qty = max(1, int(drawing.quantity or 1))
    density = _internal_density(ai_data, drawing)

    total_area_mm2 = blank_length * blank_width * product_qty
    total_weight_kg = total_area_mm2 * thickness * density

    return (
        round(total_weight_kg, 4),
        round(blank_length, 1),
        round(blank_width, 1),
    )


def _internal_sheet_material_quantity_kg(
    ai_data: dict,
    drawing: Drawing,
) -> float | None:
    basis = _internal_sheet_material_basis(ai_data, drawing)
    return basis[0] if basis else None


def _feature_quantity(items: list, default_each: int = 1) -> int:
    total = 0
    for item in items or []:
        if not isinstance(item, dict):
            continue
        try:
            qty = int(item.get("quantity") or default_each)
        except (TypeError, ValueError):
            qty = default_each
        total += max(0, qty)
    return total


def _max_drawing_dimension(ai_data: dict) -> float:
    values: list[float] = []
    for dim in ai_data.get("dimensions") or []:
        if not isinstance(dim, dict):
            continue
        try:
            value = float(dim.get("value_mm") or 0)
        except (TypeError, ValueError):
            continue
        if value > 0:
            values.append(value)
    return max(values, default=100.0)


def _is_rotational_part(ai_data: dict, drawing: Drawing) -> bool:
    text = _norm(
        f"{drawing.description} {' '.join(str(x) for x in (ai_data.get('notes') or []))}"
    )
    if any(word in text for word in [
        "SHAFT", "BUSH", "BUSHING", "PIN", "SPINDLE", "SLEEVE",
        "ROLLER", "CYLINDER", "CYLINDRICAL", "ROUND BAR", "ROUNDBAR",
    ]):
        return True

    diameter_signals = 0
    for dim in ai_data.get("dimensions") or []:
        if not isinstance(dim, dict):
            continue
        label = _norm(str(dim.get("label") or dim.get("type") or ""))
        if any(key in label for key in ["DIAMETER", "DIA", "OD", "ID"]):
            diameter_signals += 1

    return diameter_signals >= 1 and not _is_sheet_based_part(ai_data, drawing)


def _recommended_processes(ai_data: dict, drawing: Drawing) -> list[tuple[str, str, int]]:
    """Research-informed starting recommendations; engineer remains final authority."""
    recommendations: list[tuple[str, str, int]] = []
    sheet = _is_sheet_based_part(ai_data, drawing)
    rotational = _is_rotational_part(ai_data, drawing)
    holes = _feature_quantity(ai_data.get("holes") or [], 1)
    threads = _feature_quantity(ai_data.get("threads") or [], 1)
    chamfers = _feature_quantity(ai_data.get("chamfers") or [], 1)
    bends = _feature_quantity(ai_data.get("bends") or [], 1)

    if sheet:
        recommendations.append((
            "PROC-LASER",
            "Uniform-thickness sheet/plate profile: default blanking recommendation; change from dropdown if your shop uses shear/plasma/waterjet/hand cutting.",
            78,
        ))
        if bends:
            recommendations.append((
                "PROC-BEND",
                f"{bends} bend feature(s) detected.",
                90,
            ))
    elif rotational:
        recommendations.append((
            "PROC-TURN",
            "Rotational/cylindrical geometry detected; turning is the default primary machining route.",
            88,
        ))
        # Cross holes / non-rotational secondary features can require milling.
        if holes >= 2 and any(
            word in _norm(str(hole.get("type") or hole.get("callout") or ""))
            for hole in (ai_data.get("holes") or []) if isinstance(hole, dict)
            for word in ["SLOT", "CROSS", "SIDE", "OFFAXIS"]
        ):
            recommendations.append((
                "PROC-MILL",
                "Rotational body also has non-axis features; secondary milling may be required.",
                70,
            ))
    else:
        solid_feature_count = holes + threads + chamfers
        if solid_feature_count or (ai_data.get("dimensions") and not sheet):
            recommendations.append((
                "PROC-MILL",
                "Non-sheet prismatic/solid geometry with machined features: CNC milling is the default starting route.",
                72,
            ))

    if threads and not sheet:
        recommendations.append(("PROC-THREAD", f"{threads} thread feature(s) detected.", 90))
    if chamfers and not sheet:
        recommendations.append(("PROC-CHAMFER", f"{chamfers} chamfer feature(s) detected.", 85))

    return recommendations


_TIME_UNIT_HOURS = {
    "SEC": 1.0 / 3600.0,
    "SECOND": 1.0 / 3600.0,
    "SECONDS": 1.0 / 3600.0,
    "MIN": 1.0 / 60.0,
    "MINUTE": 1.0 / 60.0,
    "MINUTES": 1.0 / 60.0,
    "HR": 1.0,
    "HOUR": 1.0,
    "HOURS": 1.0,
    "SHIFT": 8.0,
    "DAY": 8.0,
}


def _time_unit_hours(unit: str) -> float | None:
    return _TIME_UNIT_HOURS.get(_norm(unit))


def _hours_to_rate_quantity(hours: float, unit: str) -> float:
    """Convert an engineering time estimate in hours into the Rate Master unit."""
    hours_per_unit = _time_unit_hours(unit)
    if hours_per_unit is None:
        return 1.0
    qty = max(0.0, float(hours or 0)) / hours_per_unit
    if _norm(unit) in {"SEC", "SECOND", "SECONDS"}:
        return round(qty)
    if _norm(unit) in {"MIN", "MINUTE", "MINUTES"}:
        return round(qty, 1)
    return round(qty, 2)


def _rate_quantity_to_hours(quantity: float, unit: str) -> float | None:
    hours_per_unit = _time_unit_hours(unit)
    if hours_per_unit is None:
        return None
    return max(0.0, float(quantity or 0)) * hours_per_unit


def _estimated_qty_for_process(
    rate_id: str | None,
    unit: str,
    ai_data: dict,
    drawing: Drawing,
) -> float:
    """
    Machine-time processes are estimated in hours internally, then converted
    to the engineer's selected Rate Master unit (sec/min/hr/shift/day).
    Fixed units such as job/setup remain quantity 1.
    """
    if _time_unit_hours(unit) is not None:
        return _hours_to_rate_quantity(
            _estimate_process_hours(rate_id, ai_data, drawing),
            unit,
        )
    return 1.0


def _estimated_time_note(quantity: float, unit: str) -> str:
    if _time_unit_hours(unit) is None:
        return ""
    return f" · estimated {quantity:g} {unit}"


def _estimate_process_hours(rate_id: str | None, ai_data: dict, drawing: Drawing) -> float:
    """
    Editable first-pass machine-hour estimate. It is intentionally deterministic,
    based on drawing size/features, and is NOT treated as a guaranteed cycle time.
    """
    max_dim = max(20.0, _max_drawing_dimension(ai_data))
    size_factor = min(4.0, max(0.5, max_dim / 100.0))
    holes = _feature_quantity(ai_data.get("holes") or [], 1)
    threads = _feature_quantity(ai_data.get("threads") or [], 1)
    chamfers = _feature_quantity(ai_data.get("chamfers") or [], 1)
    qty = max(1, int(drawing.quantity or 1))

    if rate_id == "PROC-TURN":
        hours = 0.30 + 0.18 * size_factor + 0.04 * holes + 0.08 * threads + 0.03 * chamfers
    elif rate_id == "PROC-MILL":
        hours = 0.45 + 0.24 * size_factor + 0.07 * holes + 0.09 * threads + 0.04 * chamfers
    elif rate_id == "PROC-MANUAL-MILL":
        hours = 0.55 + 0.30 * size_factor + 0.09 * holes + 0.10 * threads + 0.05 * chamfers
    elif rate_id == "PROC-MACHINE":
        hours = 0.45 + 0.25 * size_factor + 0.06 * holes + 0.08 * threads + 0.04 * chamfers
    elif rate_id == "PROC-HAND-GRIND":
        hours = 0.25 + 0.08 * size_factor
    else:
        return 1.0

    # Round to quarter-hours so the starting estimate is practical to edit.
    total = max(0.25, hours * qty)
    return round(total * 4) / 4


def _add_labour_rows(rows: list[Row]) -> None:
    process_ids = {
        row.rateId
        for row in rows
        if row.category == "PROCESS" and row.rateId
    }

    labour_added: set[str] = set()

    def add_labour(
        row_id: str,
        rate_id: str,
        item: str,
        hours: float,
        description: str,
    ):
        if rate_id in labour_added:
            return
        labour_added.add(rate_id)

        saved_labour = rate_by_id(rate_id)
        labour_unit = str(saved_labour.get("unit") or "hr") if saved_labour else "hr"
        labour_qty = (
            _hours_to_rate_quantity(hours, labour_unit)
            if _time_unit_hours(labour_unit) is not None
            else 1.0
        )

        rows.append(
            row_from_rate(
                row_id,
                "LABOUR",
                item,
                description + _estimated_time_note(labour_qty, labour_unit),
                labour_qty,
                "Estimated",
                rate_id,
                labour_unit,
                fallback_critical=70,
            )
        )

    fabrication_ids = {"PROC-LASER", "PROC-BEND", "PROC-STUD", "PROC-SAW"}
    machining_ids = {"PROC-TURN", "PROC-MILL", "PROC-MANUAL-MILL", "PROC-BORE", "PROC-THREAD", "PROC-CHAMFER", "PROC-MACHINE"}
    finishing_ids = {"PROC-GRIND", "PROC-DEBURR", "PROC-POLISH", "PROC-PASS", "PROC-COAT"}

    if process_ids & fabrication_ids:
        add_labour(
            "lab-fabricator",
            "LAB-FAB",
            "Fabricator",
            1.0,
            "Default fabrication labour allowance",
        )

    if process_ids & {"PROC-LASER", "PROC-SAW"}:
        add_labour(
            "lab-machine-operator",
            "LAB-MACHINE",
            "Machine Operator",
            0.75,
            "Default machine-operation labour allowance",
        )

    if process_ids & machining_ids:
        machine_hours = sum(
            _rate_quantity_to_hours(row.costingQty, row.unit) or 0.0
            for row in rows
            if row.category == "PROCESS"
            and row.rateId in machining_ids
        )
        machining_count = len(process_ids & machining_ids)
        estimated_labour = machine_hours if machine_hours > 0 else max(0.75, machining_count * 0.75)
        add_labour(
            "lab-machinist",
            "LAB-MACH",
            "Machinist",
            round(estimated_labour, 2),
            "Machinist time follows estimated machining hours; editable before quotation",
        )

    if "PROC-TIG" in process_ids:
        add_labour(
            "lab-tig",
            "LAB-TIG",
            "TIG Welder",
            1.0,
            "Default welding labour allowance",
        )
    elif "PROC-MIG" in process_ids:
        add_labour(
            "lab-mig",
            "LAB-MIG",
            "MIG Welder",
            1.0,
            "Default welding labour allowance",
        )
    elif "PROC-WELD" in process_ids:
        add_labour(
            "lab-welder",
            "LAB-WELD",
            "Welder / Fabricator",
            1.0,
            "Default welding labour allowance",
        )

    if process_ids & finishing_ids:
        add_labour(
            "lab-finishing",
            "LAB-FINISH",
            "Finishing Operator",
            0.75,
            "Default finishing labour allowance",
        )

    if process_ids:
        add_labour(
            "lab-qc",
            "LAB-QC",
            "QC / Handling",
            0.5,
            "Default inspection/handling labour allowance",
        )

def ai_result_to_rows(ai_data: dict, drawing: Drawing) -> list[Row]:
    """
    Convert Vision AI extraction into editable costing rows.
    Quantities come from the drawing/AI. Prices come ONLY from Rate Master.
    Missing rates are shown as 0 with Critical Score 100, rather than silently
    generating a fake quotation.
    """
    rows: list[Row] = []
    seen: set[str] = set()

    material_rate = find_material_rate_from_ai(ai_data)
    material_name = drawing.material or "Material"

    material_conf = (ai_data.get("confidence") or {}).get("material", 0)
    weight_conf = (ai_data.get("confidence") or {}).get("weight", 0)
    material_row_conf = ai_confidence_label(min(
        float(material_conf or 0),
        float(weight_conf or material_conf or 0),
    ))

    if drawing.material != "Not detected":
        internal_material_basis = _internal_sheet_material_basis(
            ai_data,
            drawing,
        )
        internal_material_qty = (
            internal_material_basis[0]
            if internal_material_basis
            else None
        )
        material_basis_note = (
            f"{max(1, drawing.quantity)} part(s) · "
            f"costing blank {internal_material_basis[1]:g} × "
            f"{internal_material_basis[2]:g} mm each "
            f"(drawing envelope + {_MATERIAL_BLANK_MARGIN_MM:g} mm)"
            if internal_material_basis
            else f"{max(1, drawing.quantity)} part(s)"
        )

        rate_id = material_rate.get("id") if material_rate else None
        fallback_source = (
            "Rate Master"
            if material_rate
            else "RATE MISSING - edit rate in this sheet"
        )

        costing_qty = (
            internal_material_qty
            if internal_material_qty is not None
            else (drawing.weight_kg if drawing.weight_kg > 0 else 0)
        )

        rows.append(
            row_from_rate(
                "ai-material-stock" if internal_material_qty is not None else "ai-material",
                "MATERIAL",
                material_name,
                material_basis_note,
                costing_qty,
                material_row_conf,
                rate_id,
                "kg",
                fallback_rate=0,
                fallback_critical=100 if not material_rate else int(material_rate.get("critical_score", 95)),
                fallback_source=fallback_source,
            )
        )

    # Manufacturing processes extracted by Vision.
    for index, process in enumerate(ai_data.get("manufacturing_processes") or [], start=1):
        if not isinstance(process, dict):
            continue

        name = str(process.get("process") or "").strip()
        if not name:
            continue

        rate_id = process_rate_id(name)
        saved_rate = rate_by_id(rate_id) if rate_id else None

        # Avoid duplicate semantic process rows.
        semantic_key = rate_id or _norm(name)
        if semantic_key in seen:
            continue
        seen.add(semantic_key)

        confidence = ai_confidence_label(process.get("confidence"))
        reason = str(process.get("reason") or "").strip()

        process_unit = saved_rate.get("unit", "job") if saved_rate else "job"
        process_qty = (
            _estimated_qty_for_process(rate_id, process_unit, ai_data, drawing)
            if saved_rate
            else 1
        )
        qty_note = (
            (reason or "Detected by Vision AI")
            + _estimated_time_note(process_qty, process_unit)
        )

        rows.append(
            row_from_rate(
                f"ai-process-{index}",
                "PROCESS",
                saved_rate.get("name") if saved_rate else name,
                qty_note,
                process_qty,
                confidence,
                rate_id if saved_rate else None,
                process_unit,
                fallback_rate=0,
                fallback_critical=100 if not saved_rate else int(saved_rate.get("critical_score", 80)),
                fallback_source="RATE MISSING - add process rate in Rate Master" if not saved_rate else "Rate Master",
            )
        )

    # Research-informed default process recommendation. The AI-extracted process
    # remains preferred; these rows only fill gaps and are editable/dropdown-selectable.
    for rec_index, (rec_rate_id, rec_reason, rec_conf) in enumerate(
        _recommended_processes(ai_data, drawing),
        start=1,
    ):
        if rec_rate_id in seen:
            continue
        saved_rate = rate_by_id(rec_rate_id)
        if not saved_rate:
            continue
        seen.add(rec_rate_id)
        unit = str(saved_rate.get("unit") or "job")
        qty = _estimated_qty_for_process(rec_rate_id, unit, ai_data, drawing)
        drawing_note = rec_reason + _estimated_time_note(qty, unit)
        rows.append(
            row_from_rate(
                f"recommended-process-{rec_index}",
                "PROCESS",
                str(saved_rate.get("name") or rec_rate_id),
                drawing_note,
                qty,
                ai_confidence_label(rec_conf),
                rec_rate_id,
                unit,
                fallback_critical=int(saved_rate.get("critical_score", 80)),
            )
        )

    # Hole features: for ordinary holes use drilling/boring quantity.
    holes = ai_data.get("holes") or []
    total_plain_holes = 0
    total_threaded_holes = 0
    hole_confidences = []

    for hole in holes:
        if not isinstance(hole, dict):
            continue
        qty = int(hole.get("quantity") or 0)
        if qty <= 0:
            continue
        hole_type = str(hole.get("type") or "")
        hole_confidences.append(hole.get("confidence"))
        if any(word in hole_type.casefold() for word in ["thread", "tap", "m16", "m12", "m10", "m8", "m6", "m5", "m4", "m3"]):
            total_threaded_holes += qty
        else:
            total_plain_holes += qty

    if total_plain_holes and "PROC-BORE" not in seen:
        seen.add("PROC-BORE")
        rows.append(
            row_from_rate(
                "ai-holes",
                "PROCESS",
                "Drilling / Boring",
                f"{total_plain_holes} hole(s)",
                1,
                ai_confidence_label(max([x or 0 for x in hole_confidences], default=70)),
                "PROC-BORE",
                "job",
                fallback_critical=85,
            )
        )

    if total_threaded_holes and "PROC-THREAD" not in seen:
        seen.add("PROC-THREAD")
        rows.append(
            row_from_rate(
                "ai-thread",
                "PROCESS",
                "Threading / Tapping",
                f"{total_threaded_holes} threaded hole(s)",
                1,
                ai_confidence_label(max([x or 0 for x in hole_confidences], default=70)),
                "PROC-THREAD",
                "job",
                fallback_critical=90,
            )
        )

    # Feature-driven processing rows, in case the AI did not explicitly list them.
    bends = ai_data.get("bends") or []
    bend_qty = sum(
        max(1, int(bend.get("quantity") or 1))
        for bend in bends
        if isinstance(bend, dict)
    )
    if bend_qty and "PROC-BEND" not in seen:
        seen.add("PROC-BEND")
        rows.append(
            row_from_rate(
                "ai-bends",
                "PROCESS",
                "Press Brake Forming",
                f"{bend_qty} bend(s)",
                1,
                "Estimated",
                "PROC-BEND",
                "job",
            )
        )

    chamfers = ai_data.get("chamfers") or []
    chamfer_qty = sum(
        max(1, int(chamfer.get("quantity") or 1))
        for chamfer in chamfers
        if isinstance(chamfer, dict)
    )
    if chamfer_qty and "PROC-CHAMFER" not in seen:
        seen.add("PROC-CHAMFER")
        rows.append(
            row_from_rate(
                "ai-chamfers",
                "PROCESS",
                "Chamfering",
                f"{chamfer_qty} chamfer location(s)",
                1,
                "Estimated",
                "PROC-CHAMFER",
                "job",
            )
        )

    # Stud welding features.
    studs = ai_data.get("studs") or []
    stud_qty = sum(
        int(stud.get("quantity") or 0)
        for stud in studs
        if isinstance(stud, dict)
    )
    if stud_qty and "PROC-STUD" not in seen:
        seen.add("PROC-STUD")
        rows.append(
            row_from_rate(
                "ai-studs",
                "PROCESS",
                "Stud Welding",
                f"{stud_qty} stud(s)",
                stud_qty,
                "Exact",
                "PROC-STUD",
                "stud",
            )
        )

    # Weld length only if AI provides an actual length.
    welds = ai_data.get("welds") or []
    total_weld_mm = 0.0
    weld_type = ""
    weld_score = 0
    for weld in welds:
        if not isinstance(weld, dict):
            continue
        try:
            total_weld_mm += float(weld.get("length_mm") or 0)
        except (TypeError, ValueError):
            pass
        weld_type = weld_type or str(weld.get("type") or "")
        try:
            weld_score = max(weld_score, int(weld.get("confidence") or 0))
        except (TypeError, ValueError):
            pass

    if total_weld_mm > 0:
        weld_rate_id = "PROC-TIG" if "TIG" in weld_type.upper() else (
            "PROC-MIG" if "MIG" in weld_type.upper() else None
        )
        rows.append(
            row_from_rate(
                "ai-weld",
                "PROCESS",
                f"{weld_type or 'Welding'}",
                f"{total_weld_mm:g} mm",
                total_weld_mm / 1000.0,
                ai_confidence_label(weld_score),
                weld_rate_id,
                "m",
                fallback_rate=0,
                fallback_critical=100,
                fallback_source="RATE / WELD TYPE REVIEW REQUIRED",
            )
        )
    elif welds and not ({"PROC-TIG", "PROC-MIG", "PROC-WELD"} & seen):
        seen.add("PROC-WELD")
        rows.append(
            row_from_rate(
                "ai-weld-job",
                "PROCESS",
                "General / Tack Welding",
                f"{len(welds)} weld callout(s)",
                1,
                "Estimated",
                "PROC-WELD",
                "job",
            )
        )

    _add_labour_rows(rows)

    return rows


def make_drawing_preview(content: bytes, filename: str) -> str:
    """
    Return a compact first-page snapshot as a data URL for the Review screen.
    This is display-only and separate from the high-resolution images used by AI.
    """
    suffix = Path(filename or "").suffix.lower()

    try:
        if suffix == ".pdf":
            doc = fitz.open(stream=content, filetype="pdf")
            if doc.page_count == 0:
                doc.close()
                return ""

            page = doc[0]

            # About 120 DPI: clear enough for a preview without sending a huge JSON response.
            matrix = fitz.Matrix(120 / 72, 120 / 72)
            pix = page.get_pixmap(matrix=matrix, alpha=False)
            data = pix.tobytes("jpeg", jpg_quality=72)
            doc.close()

            encoded = base64.b64encode(data).decode("ascii")
            return f"data:image/jpeg;base64,{encoded}"

        mime_by_suffix = {
            ".png": "image/png",
            ".jpg": "image/jpeg",
            ".jpeg": "image/jpeg",
            ".webp": "image/webp",
        }

        mime = mime_by_suffix.get(suffix)
        if mime:
            encoded = base64.b64encode(content).decode("ascii")
            return f"data:{mime};base64,{encoded}"

    except Exception:
        return ""

    return ""

def reviewed_memory(file_hash: str):
    if not load("settings", defaults_settings())["learn_from_corrections"]:
        return None
    return latest_review_by_hash(file_hash)


@app.get("/")
def root():
    return {
        "service": "dfab-quotation-api",
        "status": "ok",
        "version": "0.8.7",
    }


@app.get("/api/health")
def health():
    connected = db_ping()

    return {
        "status": "ok" if connected else "degraded",
        "version": "0.8.7",
        "database": "connected" if connected else "unavailable",
    }


@app.get("/api/database/stats")
def get_database_stats():
    if not db_ping():
        raise HTTPException(
            status_code=503,
            detail="PostgreSQL database is unavailable.",
        )

    return database_stats()


@app.get("/api/settings")
def get_settings():
    return load("settings", defaults_settings())


@app.put("/api/settings")
def put_settings(value: Settings):
    save("settings", value.model_dump())
    return value


@app.post("/api/dfm/generate")
def generate_dfm_report(value: EngineeringArtifactReq):
    return _generate_dfm(value)


@app.post("/api/bom/generate")
def generate_bom_report(value: EngineeringArtifactReq):
    return _generate_bom(value)


@app.post("/api/dfm/export")
def export_dfm_report(payload: dict):
    out = BytesIO()
    doc = SimpleDocTemplate(
        out,
        pagesize=A4,
        rightMargin=12 * mm,
        leftMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
    )
    styles = getSampleStyleSheet()

    story = [
        Paragraph(str(payload.get("name") or "DFM Report"), styles["Title"]),
        Paragraph(
            f"Drawing: {payload.get('drawing_no', '')} &nbsp;&nbsp; "
            f"Revision: {payload.get('revision', '')}",
            styles["Normal"],
        ),
        Paragraph(
            f"Manufacturing type: {payload.get('classification', '')} &nbsp;&nbsp; "
            f"Status: {payload.get('status', '')}",
            styles["Normal"],
        ),
        Spacer(1, 8),
        Paragraph("Manufacturing Feasibility", styles["Heading2"]),
    ]

    check_rows = [["Area", "Result", "Finding", "Recommendation", "Reference"]]
    for item in payload.get("checks") or []:
        check_rows.append([
            str(item.get("area", "")),
            str(item.get("result", "")),
            str(item.get("finding", "")),
            str(item.get("recommendation", "")),
            str(item.get("standard", "")),
        ])

    checks = Table(
        check_rows,
        colWidths=[28 * mm, 17 * mm, 48 * mm, 55 * mm, 34 * mm],
        repeatRows=1,
    )
    checks.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#17365d")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#c7d3df")),
    ]))
    story.extend([checks, Spacer(1, 8), Paragraph("Process Plan", styles["Heading2"])])

    plan_rows = [["Seq", "Process", "Tooling / Method", "Feasibility", "Inspection"]]
    for item in payload.get("process_plan") or []:
        plan_rows.append([
            str(item.get("sequence", "")),
            str(item.get("process", "")),
            str(item.get("tooling", "")),
            str(item.get("feasibility", "")),
            str(item.get("inspection", "")),
        ])

    plan = Table(plan_rows, colWidths=[12 * mm, 38 * mm, 60 * mm, 25 * mm, 47 * mm], repeatRows=1)
    plan.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#24568e")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#c7d3df")),
    ]))
    story.append(plan)
    doc.build(story)
    out.seek(0)

    safe = re.sub(r"[^A-Za-z0-9._-]+", "_", str(payload.get("drawing_no") or "drawing"))
    return StreamingResponse(
        out,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="DFM_{safe}.pdf"'},
    )


@app.post("/api/bom/export")
def export_bom_pdf(payload: dict):
    out = BytesIO()
    doc = SimpleDocTemplate(
        out,
        pagesize=landscape(A4),
        rightMargin=10 * mm,
        leftMargin=10 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
    )
    styles = getSampleStyleSheet()

    story = [
        Paragraph(str(payload.get("name") or "Bill of Materials"), styles["Title"]),
        Paragraph(
            f"Drawing: {payload.get('drawing_no', '')} &nbsp;&nbsp; "
            f"Revision: {payload.get('revision', '')} &nbsp;&nbsp; "
            f"Description: {payload.get('description', '')}",
            styles["Normal"],
        ),
        Spacer(1, 8),
    ]

    columns = [
        "Item",
        "Category",
        "Description",
        "Material",
        "Specification",
        "Dimensions / Size",
        "Qty",
        "Unit",
        "Weight kg",
        "Unit Cost",
        "Total Cost",
        "Remarks",
    ]

    table_rows = [columns]
    for item in payload.get("items") or []:
        table_rows.append([
            str(item.get("item_no", "")),
            str(item.get("category", "")),
            str(item.get("description", "")),
            str(item.get("material", "")),
            str(item.get("specification", "")),
            str(item.get("dimensions", "")),
            str(item.get("quantity", "")),
            str(item.get("unit", "")),
            str(item.get("weight_kg", "")),
            str(item.get("unit_cost", "")),
            str(item.get("total_cost", "")),
            str(item.get("remarks", "")),
        ])

    bom_table = Table(
        table_rows,
        colWidths=[
            12 * mm,
            24 * mm,
            34 * mm,
            25 * mm,
            28 * mm,
            32 * mm,
            15 * mm,
            15 * mm,
            20 * mm,
            22 * mm,
            22 * mm,
            28 * mm,
        ],
        repeatRows=1,
    )
    bom_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#17365d")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#c7d3df")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
    ]))
    story.append(bom_table)

    notes = payload.get("notes") or []
    if notes:
        story.extend([Spacer(1, 8), Paragraph("Notes", styles["Heading2"])])
        for note in notes:
            story.append(Paragraph(f"• {str(note)}", styles["Normal"]))

    doc.build(story)
    out.seek(0)

    safe = re.sub(r"[^A-Za-z0-9._-]+", "_", str(payload.get("drawing_no") or "drawing"))
    return StreamingResponse(
        out,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="BOM_{safe}.pdf"'},
    )


@app.get("/api/rates", response_model=list[RateItem])
def get_rates():
    repaired, _ = ensure_default_rates()
    return repaired


@app.post("/api/rates/restore-defaults", response_model=list[RateItem])
def restore_default_rates():
    """
    Restore all built-in starter rows and starter values where a built-in
    row is missing/zero. Positive engineer-entered rates and custom rows stay unchanged.
    """
    repaired, _ = ensure_default_rates()
    return repaired


@app.get("/api/rate-catalog")
def rate_catalog():
    rows = rates_list()

    materials: dict[str, list[str]] = {}
    for row in rows:
        if row.get("category") == "MATERIAL":
            material_name = str(row.get("name") or "Material")
            materials.setdefault(material_name, [])
            grade = str(row.get("grade") or "")
            if grade and grade not in materials[material_name]:
                materials[material_name].append(grade)

    base_processes = [
        "Laser Cutting",
        "Shearing",
        "Plasma Cutting",
        "Waterjet Cutting",
        "Press Brake Forming",
        "CNC Turning",
        "CNC Milling",
        "Manual Milling",
        "General Machining",
        "Saw / Raw Stock Cutting",
        "Drilling / Boring",
        "Threading / Tapping",
        "Chamfering",
        "TIG Welding",
        "MIG Welding",
        "General / Tack Welding",
        "Stud Welding",
        "Grinding & Flush",
        "Hand Grinding / Cut-off",
        "Deburring",
        "Polishing",
        "Passivation",
        "Powder Coating",
        "Inspection & Handling",
    ]

    base_labour = [
        "Fabricator",
        "Machinist",
        "Machine Operator",
        "TIG Welder",
        "MIG Welder",
        "Welder / Fabricator",
        "Finishing Operator",
        "QC / Handling",
    ]

    # sec/min/hr/job are deliberately explicit because the selected Rate Master
    # unit directly controls how auto-estimated quantities are converted/costed.
    base_units = [
        "sec",
        "min",
        "hr",
        "shift",
        "day",
        "job",
        "setup",
        "each",
        "piece",
        "kg",
        "g",
        "ton",
        "mm",
        "m",
        "m2",
        "m3",
        "hole",
        "stud",
        "bend",
        "cut",
        "weld",
        "%",
    ]

    def unique(values: list[str]) -> list[str]:
        seen: set[str] = set()
        result: list[str] = []
        for value in values:
            clean = str(value or "").strip()
            key = clean.casefold()
            if clean and key not in seen:
                seen.add(key)
                result.append(clean)
        return result

    saved_processes = [
        str(row.get("name") or "")
        for row in rows
        if row.get("category") == "PROCESS"
    ]
    saved_labour = [
        str(row.get("name") or "")
        for row in rows
        if row.get("category") == "LABOUR"
    ]
    saved_units = [str(row.get("unit") or "") for row in rows]

    return {
        "materials": materials,
        "processes": unique(base_processes + saved_processes),
        "labour": unique(base_labour + saved_labour),
        "commercial": ["Material Wastage", "Factory Overhead", "Profit / Markup"],
        "units": unique(base_units + saved_units),
    }


def _normalized_rate_category(value: str) -> str:
    key = _norm(value)
    return {
        "MATERIAL": "MATERIAL",
        "MATERIALS": "MATERIAL",
        "PROCESS": "PROCESS",
        "PROCESSING": "PROCESS",
        "OPERATION": "PROCESS",
        "LABOUR": "LABOUR",
        "LABOR": "LABOUR",
        "OTHER": "OTHER",
        "COMMERCIAL": "COMMERCIAL",
    }.get(key, "OTHER")


def _default_rate_critical(category: str) -> int:
    return {
        "MATERIAL": 85,
        "PROCESS": 70,
        "LABOUR": 65,
        "COMMERCIAL": 75,
        "OTHER": 50,
    }.get(category, 50)


def _find_rate_for_sync(
    category: str,
    name: str,
    grade: str,
) -> dict | None:
    for rate in rates_list():
        if not rate.get("active", True):
            continue
        if rate.get("category") != category:
            continue
        if str(rate.get("name", "")).casefold().strip() != name.casefold().strip():
            continue

        if category == "MATERIAL":
            saved_grade = str(rate.get("grade", "")).casefold().strip()
            if saved_grade != grade.casefold().strip():
                continue

        return rate

    return None


@app.post("/api/rates/sync-row")
def sync_rate_from_cost_sheet(value: RateSyncReq):
    row = value.row
    category = _normalized_rate_category(row.category)

    if category == "MATERIAL":
        name = value.material_family.strip() or row.item.strip() or "Material"
        grade = " / ".join(
            part
            for part in [
                value.material_grade.strip(),
                value.material_specification.strip(),
            ]
            if part
        )
    else:
        name = row.item.strip() or "Unnamed Cost Item"
        grade = ""

    existing = None

    if row.rateId:
        candidate = rate_by_id(row.rateId)

        if candidate:
            same_category = candidate.get("category") == category
            same_name = str(candidate.get("name", "")).casefold().strip() == name.casefold().strip()
            same_grade = (
                category != "MATERIAL"
                or str(candidate.get("grade", "")).casefold().strip() == grade.casefold().strip()
            )

            if same_category and same_name and same_grade:
                existing = candidate

    if existing is None:
        existing = _find_rate_for_sync(category, name, grade)

    all_rates = rates_list()

    if existing:
        rate_id = str(existing["id"])
        index = next(
            i
            for i, saved in enumerate(all_rates)
            if saved.get("id") == rate_id
        )

        saved = all_rates[index]

        item = {
            **saved,
            "category": category,
            "name": name,
            "grade": grade,
            "unit": row.unit.strip() or saved.get("unit", "job"),
            "price": max(0.0, float(row.rate or 0)),
            "active": True,
            "updated_at": now(),
        }

        all_rates[index] = item
    else:
        rate_id = "RATE-AUTO-" + uuid.uuid4().hex[:10].upper()

        item = {
            "id": rate_id,
            "category": category,
            "name": name,
            "grade": grade,
            "unit": row.unit.strip() or "job",
            "price": max(0.0, float(row.rate or 0)),
            "critical_score": _default_rate_critical(category),
            "active": True,
            "notes": "Auto-added from costing sheet",
            "updated_at": now(),
        }

        all_rates.append(item)

    save("rates", all_rates)

    row.category = category
    row.rateId = rate_id
    row.rateSource = "Rate Master"
    row.criticalScore = int(
        item.get("critical_score", _default_rate_critical(category))
    )
    row.rate = float(item.get("price", row.rate))
    row.cost = row.costingQty * row.rate

    return {
        "row": row,
        "rate": item,
    }


@app.post("/api/rates", response_model=RateItem)
def add_rate(value: RateItem):
    rows = rates_list()
    item = value.model_dump()
    item["id"] = item.get("id") or "RATE-" + uuid.uuid4().hex[:10].upper()
    if any(r.get("id") == item["id"] for r in rows):
        item["id"] = "RATE-" + uuid.uuid4().hex[:10].upper()
    item["updated_at"] = now()
    rows.append(item)
    save("rates", rows)
    return item


@app.put("/api/rates/{rate_id}", response_model=RateItem)
def update_rate(rate_id: str, value: RateItem):
    rows = rates_list()
    index = next((i for i, r in enumerate(rows) if r.get("id") == rate_id), None)
    if index is None:
        raise HTTPException(status_code=404, detail="Rate not found")
    item = value.model_dump()
    item["id"] = rate_id
    item["updated_at"] = now()
    rows[index] = item
    save("rates", rows)
    return item


@app.delete("/api/rates/{rate_id}")
def delete_rate(rate_id: str):
    rows = rates_list()
    new_rows = [r for r in rows if r.get("id") != rate_id]
    if len(new_rows) == len(rows):
        raise HTTPException(status_code=404, detail="Rate not found")
    save("rates", new_rows)
    return {"status": "deleted", "id": rate_id}


@app.post("/api/rates/apply", response_model=list[Row])
def apply_saved_rates(value: CostReq):
    updated: list[Row] = []

    for row in value.rows:
        saved = rate_by_id(row.rateId)

        if saved:
            old_unit = str(row.unit or "")
            new_unit = str(saved.get("unit") or old_unit)

            # Preserve the same predicted duration when the engineer changes
            # hr -> min -> sec (or back) in Rate Master.
            old_hours = _rate_quantity_to_hours(row.costingQty, old_unit)
            if old_hours is not None and _time_unit_hours(new_unit) is not None:
                row.costingQty = _hours_to_rate_quantity(old_hours, new_unit)
            elif _time_unit_hours(new_unit) is None and old_hours is not None:
                # Switching a timed rate to a fixed per-job/per-setup unit.
                row.costingQty = 1.0

            row.category = str(saved.get("category") or row.category)
            saved_name = str(saved.get("name") or row.item)
            saved_grade = str(saved.get("grade") or "")

            row.item = (
                f"{saved_name} — {saved_grade}"
                if row.category == "MATERIAL" and saved_grade
                else saved_name
            )
            row.rate = float(saved.get("price", row.rate))
            row.unit = new_unit
            row.criticalScore = int(saved.get("critical_score", row.criticalScore))
            row.rateSource = "Rate Master"
            row.cost = row.costingQty * row.rate

        updated.append(row)

    return updated


@app.post("/api/analyze")
async def analyze(file: UploadFile = File(...), force_ai: bool = True):
    content = await file.read()

    if not content:
        raise HTTPException(status_code=400, detail="Uploaded file is empty.")

    filename = file.filename or "drawing.pdf"
    extension = Path(filename).suffix.lower()
    # Fast path: do not render/base64 a preview inside the analysis request.
    # The browser displays the original uploaded PDF/image directly via ObjectURL.
    preview_image = ""

    file_hash = sha256(content).hexdigest()
    extraction_id = "EXT-" + uuid.uuid4().hex[:10].upper()

    # Reuse corrected data ONLY for the exact same file hash.
    memory = None if force_ai else reviewed_memory(file_hash)

    ai_raw = None
    text_preview = ""

    if memory:
        drawing = Drawing(**memory["drawing"])
        rows = [Row(**x) for x in memory["rows"]]
        source = "reviewed_memory"
        warnings = [
            "Exact previously reviewed drawing found. Saved engineer corrections were reused."
        ]

    elif extension == ".pdf":
        try:
            # PDF -> 300 DPI page image -> Vision AI -> structured JSON
            ai_raw = analyze_pdf_with_ai(content)

            if not isinstance(ai_raw, dict):
                raise ValueError("Vision AI did not return a JSON object.")

            drawing = ai_result_to_drawing(ai_raw)

            # Build editable costing rows from the AI extraction.
            # Quantities/features come from the drawing; prices come from Rate Master.
            rows = ai_result_to_rows(ai_raw, drawing)

            source = "vision_ai"
            warnings = []

            uncertain = ai_raw.get("missing_or_uncertain") or []
            if uncertain:
                warnings.append(
                    "Vision extraction completed, but some values require review: "
                    + ", ".join(str(x) for x in uncertain[:12])
                )

            missing_rate_rows = [
                row.item
                for row in rows
                if row.rate <= 0 and row.category in {"MATERIAL", "PROCESS", "LABOUR", "OTHER"}
            ]
            if missing_rate_rows:
                warnings.append(
                    "Rate Master values are missing for: "
                    + ", ".join(missing_rate_rows[:12])
                    + ". Enter approved rates before generating the final quotation."
                )

        except Exception as exc:
            # Do not silently fall back to a weak parser. Return the real
            # upstream cause so the UI can distinguish rate-limit/transient
            # Gemini failures from drawing/schema problems.
            message = str(exc)
            upper = message.upper()

            if any(
                marker in upper
                for marker in (
                    "429",
                    "RESOURCE_EXHAUSTED",
                    "RATE LIMIT",
                )
            ):
                status_code = 429
                reason = (
                    "Gemini rate limit reached. "
                    "Wait a few seconds and retry this drawing."
                )
            elif any(
                marker in upper
                for marker in (
                    "503",
                    "UNAVAILABLE",
                    "TIMEOUT",
                    "DEADLINE_EXCEEDED",
                )
            ):
                status_code = 503
                reason = (
                    "Gemini is temporarily unavailable. "
                    "Retry this drawing shortly."
                )
            else:
                status_code = 502
                reason = "Vision AI extraction failed."

            print(
                "[ANALYZE ERROR]",
                type(exc).__name__,
                message,
                flush=True,
            )

            raise HTTPException(
                status_code=status_code,
                detail=(
                    f"{reason} "
                    f"{type(exc).__name__}: {message}"
                ),
            ) from exc

    else:
        # Existing non-PDF safe parser/blank behavior.
        drawing, rows, source, warnings, text_preview = extract_drawing_fields(
            content, filename
        )

    settings = load("settings", defaults_settings())

    if settings.get("auto_dataset_capture", True):
        append_extraction_record(
            {
                "id": extraction_id,
                "created_at": now(),
                "filename": filename,
                "file_hash": file_hash,
                "source": source,
                "warnings": warnings,
                "drawing": drawing.model_dump(),
                "rows": [row.model_dump() for row in rows],
                "ai_raw": ai_raw,
            }
        )

    summary = calc(rows)

    artifact_payload = EngineeringArtifactReq(
        file_hash=file_hash,
        filename=filename,
        drawing=drawing,
        rows=rows,
        ai_raw=ai_raw or {},
    )

    # These are deterministic/local and very fast. Returning them with the
    # analysis avoids two additional Vercel function calls per drawing.
    dfm_report = _generate_dfm(artifact_payload)
    bom_report = _generate_bom(artifact_payload)

    return {
        "extraction_id": extraction_id,
        "file_hash": file_hash,
        "learning_source": source,
        "extraction_warnings": warnings,
        "text_preview": text_preview,
        "preview_image": preview_image,
        "drawing": drawing,
        "rows": rows,
        "summary": summary,
        "dfm": dfm_report,
        "bom": bom_report,
        "ai_raw": ai_raw,
    }


@app.post("/api/review")
def review(value: ReviewReq):
    record = {
        "id": "REVW-" + uuid.uuid4().hex[:10].upper(),
        "created_at": now(),
        "extraction_id": value.extraction_id,
        "file_hash": value.file_hash,
        "drawing": value.drawing.model_dump(),
        "rows": [r.model_dump() for r in value.rows],
    }

    reviewed_samples = append_review_record(record)

    return {
        "status": "saved",
        "reviewed_samples": reviewed_samples,
    }


@app.post("/api/calculate", response_model=Summary)
def calculate(value: CostReq):
    return calc(
        value.rows,
        material_wastage_pct=value.material_wastage_pct,
        overhead_pct=value.overhead_pct,
        markup_pct=value.markup_pct,
        material_wastage_override=value.material_wastage_override,
        overhead_override=value.overhead_override,
        markup_override=value.markup_override,
        selling_price_override=value.selling_price_override,
    )


@app.get("/api/dataset/stats")
def dataset_stats():
    counts = dataset_counts_fast()
    settings = load("settings", defaults_settings())
    meta = load(
        "dataset_meta",
        {
            "version": 1,
            "reviewed_at_version": 0,
            "training_at_version": 0,
        },
    )

    reviewed = counts["reviews"]
    training_samples = counts.get("training_samples", 0)

    new_reviewed = max(
        0,
        reviewed - meta.get("reviewed_at_version", 0),
    )
    new_training = max(
        0,
        training_samples - meta.get("training_at_version", 0),
    )

    threshold = max(
        1,
        int(settings["training_batch_threshold"]),
    )

    return {
        "extractions": counts["extractions"],
        "reviewed_samples": reviewed,
        "training_samples": training_samples,
        "unique_files": counts["unique_files"],
        "dataset_version": meta.get("version", 1),
        "new_reviewed_since_version": new_reviewed,
        "new_training_since_version": new_training,
        "training_batch_threshold": threshold,
        "batch_ready": new_training >= threshold,
        "auto_dataset_capture": settings["auto_dataset_capture"],
        "learn_from_corrections": settings["learn_from_corrections"],
    }


@app.post("/api/dataset/training")
async def add_training_sample(
    file: UploadFile = File(...),
    extraction_id: str = Form(""),
    file_hash: str = Form(""),
    customer: str = Form(""),
    drawing_json: str = Form("{}"),
    rows_json: str = Form("[]"),
    summary_json: str = Form("{}"),
    ai_raw_json: str = Form("{}"),
):
    """
    Add a quotation-approved drawing to the curated Training Dataset.
    This endpoint is called ONLY after the user explicitly clicks
    'Send to Training' after quotation download.
    """
    content = await file.read()

    if not content:
        raise HTTPException(status_code=400, detail="Training drawing file is empty.")

    actual_hash = sha256(content).hexdigest()

    try:
        drawing = json.loads(drawing_json or "{}")
        rows = json.loads(rows_json or "[]")
        summary = json.loads(summary_json or "{}")
        ai_raw = json.loads(ai_raw_json or "{}")
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=400, detail="Invalid training metadata JSON.") from exc

    # The uploaded original file is authoritative.
    final_hash = actual_hash
    if file_hash and file_hash != actual_hash:
        # Do not reject: browser/file metadata may come from an earlier upload
        # object, but record the actual bytes hash as the dataset key.
        pass

    record = {
        "id": f"TRN-{final_hash[:16].upper()}",
        "created_at": now(),
        "filename": Path(file.filename or "drawing.bin").name,
        "content_type": file.content_type or "application/octet-stream",
        "file_hash": final_hash,
        "source_file_hash": file_hash,
        "extraction_id": extraction_id,
        "customer": customer,
        "drawing": drawing,
        "rows": rows,
        "summary": summary,
        "ai_raw": ai_raw,
        "file_size": len(content),
        "training_source": "quotation_download_confirmation",
    }

    _, total = upsert_training_sample(record, content)

    return {
        "status": "saved",
        "id": record["id"],
        "drawing_no": str(drawing.get("drawing_no") or ""),
        "training_samples": total,
    }


@app.get("/api/dataset/export")
def export_dataset():
    """
    Export ONLY explicitly approved training samples.
    ZIP contains original drawing files plus JSONL target data.
    """
    samples = training_samples_for_export()
    out = BytesIO()

    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as archive:
        jsonl_lines: list[str] = []

        for index, (payload, file_content, filename) in enumerate(samples, 1):
            safe_name = re.sub(
                r"[^A-Za-z0-9._-]+",
                "_",
                Path(filename or f"drawing_{index}.bin").name,
            ).strip("._") or f"drawing_{index}.bin"

            sample_id = str(payload.get("id") or f"sample_{index}")
            archive_name = f"drawings/{index:04d}_{sample_id}_{safe_name}"

            if file_content:
                archive.writestr(archive_name, file_content)

            jsonl_lines.append(
                json.dumps(
                    {
                        "input": {
                            "training_id": sample_id,
                            "file_hash": payload.get("file_hash", ""),
                            "original_file": archive_name,
                            "filename": payload.get("filename", filename),
                            "content_type": payload.get("content_type", ""),
                            "extraction_id": payload.get("extraction_id", ""),
                            "ai_raw": payload.get("ai_raw", {}),
                        },
                        "target": {
                            "drawing": payload.get("drawing", {}),
                            "rows": payload.get("rows", []),
                            "summary": payload.get("summary", {}),
                            "customer": payload.get("customer", ""),
                        },
                    },
                    ensure_ascii=False,
                )
            )

        archive.writestr(
            "training_dataset.jsonl",
            ("\n".join(jsonl_lines) + ("\n" if jsonl_lines else "")).encode("utf-8"),
        )

    out.seek(0)

    return StreamingResponse(
        out,
        media_type="application/zip",
        headers={
            "Content-Disposition": 'attachment; filename="quotation_training_dataset.zip"'
        },
    )


@app.post("/api/revisions")
def add_revision(value: RevisionReq):
    record = {
        "id": "R-" + uuid.uuid4().hex[:8].upper(),
        "created_at": now(),
        "drawing_no": value.drawing.drawing_no,
        "revision": value.drawing.revision,
        "description": value.drawing.description,
        "material": value.drawing.material,
        "note": value.note,
    }
    rows = load("revisions", [])
    rows.append(record)
    save("revisions", rows)
    return record


@app.get("/api/revisions/{drawing_no}")
def revisions(drawing_no: str):
    return [x for x in load("revisions", []) if x.get("drawing_no") == drawing_no]


@app.post("/api/quotations")
def add_quote(value: QuoteReq):
    record = {
        "id": "Q-" + datetime.now().strftime("%Y%m%d") + "-" + uuid.uuid4().hex[:5].upper(),
        "created_at": now(),
        "customer": value.customer,
        "name": (
            f"{value.drawing.drawing_no} - {value.customer}"
            if value.drawing.drawing_no
            else f"Quotation - {value.customer}"
        ),
        "drawing_no": value.drawing.drawing_no,
        "revision": value.drawing.revision,
        "description": value.drawing.description,
        "selling_price": value.summary.selling_price,
        "status": value.status,
        "drawing": value.drawing.model_dump(),
        "summary": value.summary.model_dump(),
        "rows": [row.model_dump() for row in value.rows],
    }
    rows = load("quotations", [])
    rows.append(record)
    save("quotations", rows)
    return record


@app.post("/api/quotations/batch")
def add_batch_quote(value: BatchSaveReq):
    if not value.items:
        raise HTTPException(status_code=400, detail="No drawings supplied.")

    saved = load("quotations", [])
    created: list[dict] = []

    if value.mode == "separate":
        for item in value.items:
            record = {
                "id": "Q-" + datetime.now().strftime("%Y%m%d") + "-" + uuid.uuid4().hex[:5].upper(),
                "created_at": now(),
                "customer": value.customer,
                "name": (
                    f"{item.drawing.drawing_no} - {value.customer}"
                    if item.drawing.drawing_no
                    else f"Quotation - {value.customer}"
                ),
                "drawing_no": item.drawing.drawing_no,
                "revision": item.drawing.revision,
                "description": item.drawing.description,
                "selling_price": item.summary.selling_price,
                "status": value.status,
                "batch_mode": "separate",
                "drawing": item.drawing.model_dump(),
                "summary": item.summary.model_dump(),
                "rows": [row.model_dump() for row in item.rows],
            }
            saved.append(record)
            created.append(record)
    else:
        total = round(sum(item.summary.selling_price for item in value.items), 2)
        drawing_numbers = [item.drawing.drawing_no for item in value.items]

        record = {
            "id": "Q-" + datetime.now().strftime("%Y%m%d") + "-" + uuid.uuid4().hex[:5].upper(),
            "created_at": now(),
            "customer": value.customer,
            "name": f"Merged {len(value.items)} Drawings - {value.customer}",
            "drawing_no": f"{len(value.items)} DRAWINGS",
            "revision": "MULTI",
            "description": "Merged quotation: " + ", ".join(drawing_numbers[:8]),
            "selling_price": total,
            "status": value.status,
            "batch_mode": "merge",
            "drawing_numbers": drawing_numbers,
            "drawing_count": len(value.items),
            "items": [
                {
                    "drawing": item.drawing.model_dump(),
                    "summary": item.summary.model_dump(),
                    "rows": [row.model_dump() for row in item.rows],
                }
                for item in value.items
            ],
        }
        saved.append(record)
        created.append(record)

    save("quotations", saved)

    return {
        "mode": value.mode,
        "count": len(created),
        "records": created,
        "id": created[0]["id"],
    }


@app.get("/api/quotations")
def quotations():
    return load("quotations", [])


@app.put("/api/quotations/{quote_id}/rename")
def rename_quotation(quote_id: str, value: QuoteRenameReq):
    rows = load("quotations", [])
    clean_name = value.name.strip()

    for row in rows:
        if row.get("id") == quote_id:
            row["name"] = clean_name
            row["updated_at"] = now()
            save("quotations", rows)
            return row

    raise HTTPException(status_code=404, detail="Quotation not found.")


@app.delete("/api/quotations/{quote_id}")
def delete_quotation(quote_id: str):
    rows = load("quotations", [])
    next_rows = [
        row
        for row in rows
        if row.get("id") != quote_id
    ]

    if len(next_rows) == len(rows):
        raise HTTPException(status_code=404, detail="Quotation not found.")

    save("quotations", next_rows)
    return {
        "status": "deleted",
        "id": quote_id,
    }


@app.post("/api/export/excel")
def excel(value: ExportReq):
    wb = Workbook()
    ws = wb.active
    ws.title = "Cost Estimate"

    navy = "17365D"
    thin = Side(style="thin", color="D9D9D9")
    ws.merge_cells("A1:J1")
    ws["A1"] = "AI MANUFACTURING COST ESTIMATE"
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=16)
    ws["A1"].fill = PatternFill("solid", fgColor=navy)
    ws["A1"].alignment = Alignment(horizontal="center")

    meta = [
        ("Drawing No.", value.drawing.drawing_no),
        ("Revision", value.drawing.revision),
        ("Description", value.drawing.description),
        ("Material", value.drawing.material),
        ("Thickness mm", value.drawing.thickness_mm),
        ("Weight kg", value.drawing.weight_kg),
        ("Quantity", value.drawing.quantity),
    ]
    for i, (key, val) in enumerate(meta, 3):
        ws[f"A{i}"] = key
        ws[f"B{i}"] = val
        ws[f"A{i}"].font = Font(bold=True)

    start = 12
    headers = [
        "Category",
        "Item / Process",
        "Drawing Qty",
        "Costing Qty",
        "Unit",
        "Rate",
        "Rate Source",
        "Criticality Score",
        "Status",
        "Estimated Cost",
    ]
    for column, header in enumerate(headers, 1):
        cell = ws.cell(start, column, header)
        cell.font = Font(bold=True, color=navy)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
        cell.alignment = Alignment(horizontal="center")

    settings = load("settings", defaults_settings())
    medium = int(settings.get("critical_medium_threshold", 40))
    high = int(settings.get("critical_high_threshold", 70))

    for row_no, row in enumerate(value.rows, start + 1):
        vals = [
            row.category,
            row.item,
            row.drawingQty,
            row.costingQty,
            row.unit,
            row.rate,
            row.rateSource,
            row.criticalScore,
            row.confidence,
            row.costingQty * row.rate,
        ]
        for col_no, val in enumerate(vals, 1):
            cell = ws.cell(row_no, col_no, val)
            cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
        ws.cell(row_no, 6).number_format = "₹#,##0.00"
        ws.cell(row_no, 10).number_format = "₹#,##0.00"

        score_cell = ws.cell(row_no, 8)
        if row.criticalScore >= high:
            score_cell.fill = PatternFill("solid", fgColor="F4CCCC")
            score_cell.font = Font(color="9C0006", bold=True)
        elif row.criticalScore >= medium:
            score_cell.fill = PatternFill("solid", fgColor="FFF2CC")
            score_cell.font = Font(color="7F6000", bold=True)
        else:
            score_cell.fill = PatternFill("solid", fgColor="D9EAD3")
            score_cell.font = Font(color="274E13", bold=True)

    summary_row = start + len(value.rows) + 3
    summary = [
        ("Direct Cost", value.summary.direct_cost),
        (f"Material Wastage ({value.summary.material_wastage_pct:g}% | C{value.summary.material_wastage_critical})", value.summary.material_wastage),
        (f"Factory Overhead ({value.summary.overhead_pct:g}% | C{value.summary.overhead_critical})", value.summary.overhead),
        ("Manufacturing Cost", value.summary.manufacturing_cost),
        (f"Markup ({value.summary.markup_pct:g}% | C{value.summary.markup_critical})", value.summary.markup),
        ("Selling Price", value.summary.selling_price),
    ]
    for i, (key, val) in enumerate(summary, summary_row):
        ws[f"I{i}"] = key
        ws[f"J{i}"] = val
        ws[f"I{i}"].font = Font(bold=True)
        ws[f"J{i}"].number_format = "₹#,##0.00"

    widths = {
        "A": 15,
        "B": 38,
        "C": 18,
        "D": 14,
        "E": 12,
        "F": 14,
        "G": 20,
        "H": 18,
        "I": 15,
        "J": 18,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{value.drawing.drawing_no}_cost_estimate.xlsx"'
        },
    )


@app.post("/api/export/pdf")
def pdf(value: PdfReq):
    # Customer-facing PDF intentionally excludes internal criticality/risk scores.
    out = BytesIO()
    doc = SimpleDocTemplate(
        out,
        pagesize=A4,
        rightMargin=18 * mm,
        leftMargin=18 * mm,
        topMargin=18 * mm,
        bottomMargin=18 * mm,
    )
    styles = getSampleStyleSheet()
    story = [
        Paragraph("QUOTATION", styles["Title"]),
        Spacer(1, 5 * mm),
        Paragraph(f"<b>Customer:</b> {value.customer}", styles["BodyText"]),
        Paragraph(
            f"<b>Drawing:</b> {value.drawing.drawing_no} / {value.drawing.revision}",
            styles["BodyText"],
        ),
        Paragraph(f"<b>Description:</b> {value.drawing.description}", styles["BodyText"]),
        Spacer(1, 5 * mm),
    ]

    data = [["Item / Process", "Qty", "Unit", "Rate (INR)", "Cost (INR)"]] + [
        [
            row.item,
            f"{row.costingQty:g}",
            row.unit,
            f"{row.rate:,.2f}",
            f"{row.costingQty * row.rate:,.2f}",
        ]
        for row in value.rows
    ]
    table = Table(data, colWidths=[76 * mm, 20 * mm, 18 * mm, 26 * mm, 30 * mm], repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#17365D")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D9D9D9")),
            ]
        )
    )
    story += [table, Spacer(1, 6 * mm)]

    sums = [
        ["Direct Cost", f"INR {value.summary.direct_cost:,.2f}"],
        [f"Material Wastage ({value.summary.material_wastage_pct:g}%)", f"INR {value.summary.material_wastage:,.2f}"],
        [f"Factory Overhead ({value.summary.overhead_pct:g}%)", f"INR {value.summary.overhead:,.2f}"],
        ["Manufacturing Cost", f"INR {value.summary.manufacturing_cost:,.2f}"],
        [f"Markup ({value.summary.markup_pct:g}%)", f"INR {value.summary.markup:,.2f}"],
        ["Selling Price", f"INR {value.summary.selling_price:,.2f}"],
    ]
    summary_table = Table(sums, colWidths=[65 * mm, 45 * mm], hAlign="RIGHT")
    summary_table.setStyle(
        TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D9D9D9")),
                ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#E2F0D9")),
                ("ALIGN", (1, 0), (1, -1), "RIGHT"),
            ]
        )
    )
    story.append(summary_table)
    doc.build(story)
    out.seek(0)
    return StreamingResponse(
        out,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="{value.drawing.drawing_no}_quotation.pdf"'
        },
    )


def _safe_pdf_name(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", value or "drawing")
    return cleaned.strip("._") or "drawing"


def _render_single_batch_pdf(
    customer: str,
    item: BatchQuoteItem,
) -> bytes:
    out = BytesIO()
    doc = SimpleDocTemplate(
        out,
        pagesize=A4,
        rightMargin=18 * mm,
        leftMargin=18 * mm,
        topMargin=18 * mm,
        bottomMargin=18 * mm,
    )

    styles = getSampleStyleSheet()

    story = [
        Paragraph("QUOTATION", styles["Title"]),
        Spacer(1, 5 * mm),
        Paragraph(f"<b>Customer:</b> {customer}", styles["BodyText"]),
        Paragraph(
            f"<b>Drawing:</b> {item.drawing.drawing_no} / {item.drawing.revision}",
            styles["BodyText"],
        ),
        Paragraph(
            f"<b>Description:</b> {item.drawing.description}",
            styles["BodyText"],
        ),
        Spacer(1, 5 * mm),
    ]

    data = [["Item / Process", "Qty", "Unit", "Rate (INR)", "Cost (INR)"]] + [
        [
            row.item,
            f"{row.costingQty:g}",
            row.unit,
            f"{row.rate:,.2f}",
            f"{row.costingQty * row.rate:,.2f}",
        ]
        for row in item.rows
    ]

    table = Table(
        data,
        colWidths=[76 * mm, 20 * mm, 18 * mm, 26 * mm, 30 * mm],
        repeatRows=1,
    )
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#17365D")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D9D9D9")),
            ]
        )
    )

    story += [table, Spacer(1, 6 * mm)]

    sums = [
        ["Direct Cost", f"INR {item.summary.direct_cost:,.2f}"],
        [
            f"Material Wastage ({item.summary.material_wastage_pct:g}%)",
            f"INR {item.summary.material_wastage:,.2f}",
        ],
        [
            f"Factory Overhead ({item.summary.overhead_pct:g}%)",
            f"INR {item.summary.overhead:,.2f}",
        ],
        ["Manufacturing Cost", f"INR {item.summary.manufacturing_cost:,.2f}"],
        [
            f"Markup ({item.summary.markup_pct:g}%)",
            f"INR {item.summary.markup:,.2f}",
        ],
        ["Selling Price", f"INR {item.summary.selling_price:,.2f}"],
    ]

    summary_table = Table(
        sums,
        colWidths=[65 * mm, 45 * mm],
        hAlign="RIGHT",
    )
    summary_table.setStyle(
        TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D9D9D9")),
                ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#E2F0D9")),
                ("ALIGN", (1, 0), (1, -1), "RIGHT"),
            ]
        )
    )

    story.append(summary_table)
    doc.build(story)

    return out.getvalue()


def _render_merged_batch_pdf(
    customer: str,
    items: list[BatchQuoteItem],
) -> bytes:
    out = BytesIO()
    doc = SimpleDocTemplate(
        out,
        pagesize=A4,
        rightMargin=14 * mm,
        leftMargin=14 * mm,
        topMargin=16 * mm,
        bottomMargin=16 * mm,
    )

    styles = getSampleStyleSheet()

    story = [
        Paragraph("QUOTATION", styles["Title"]),
        Spacer(1, 4 * mm),
        Paragraph(f"<b>Customer:</b> {customer}", styles["BodyText"]),
        Paragraph(
            f"<b>Drawings:</b> {len(items)}",
            styles["BodyText"],
        ),
        Spacer(1, 5 * mm),
    ]

    data = [
        [
            "Sl.",
            "Drawing No.",
            "Rev",
            "Description",
            "Qty",
            "Unit Price (INR)",
            "Line Total (INR)",
        ]
    ]

    grand_total = 0.0

    for index, item in enumerate(items, 1):
        qty = max(1, int(item.drawing.quantity or 1))
        line_total = float(item.summary.selling_price)
        unit_price = line_total / qty if qty else line_total
        grand_total += line_total

        data.append(
            [
                str(index),
                item.drawing.drawing_no,
                item.drawing.revision,
                item.drawing.description,
                str(qty),
                f"{unit_price:,.2f}",
                f"{line_total:,.2f}",
            ]
        )

    table = Table(
        data,
        colWidths=[
            9 * mm,
            24 * mm,
            13 * mm,
            65 * mm,
            12 * mm,
            27 * mm,
            29 * mm,
        ],
        repeatRows=1,
    )

    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#17365D")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 7.4),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D9D9D9")),
                ("ALIGN", (4, 1), (-1, -1), "RIGHT"),
            ]
        )
    )

    story += [table, Spacer(1, 6 * mm)]

    total_table = Table(
        [
            ["Grand Total", f"INR {grand_total:,.2f}"],
        ],
        colWidths=[65 * mm, 45 * mm],
        hAlign="RIGHT",
    )

    total_table.setStyle(
        TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#B8C5D3")),
                ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#E2F0D9")),
                ("ALIGN", (1, 0), (1, 0), "RIGHT"),
            ]
        )
    )

    story.append(total_table)
    doc.build(story)

    return out.getvalue()


@app.post("/api/export/pdf-batch")
def batch_pdf(value: BatchPdfReq):
    if not value.items:
        raise HTTPException(status_code=400, detail="No drawings supplied.")

    if value.mode == "merge":
        payload = _render_merged_batch_pdf(
            value.customer,
            value.items,
        )

        out = BytesIO(payload)

        return StreamingResponse(
            out,
            media_type="application/pdf",
            headers={
                "Content-Disposition": 'attachment; filename="merged_quotation.pdf"'
            },
        )

    archive = BytesIO()

    with zipfile.ZipFile(
        archive,
        "w",
        compression=zipfile.ZIP_DEFLATED,
    ) as bundle:
        for index, item in enumerate(value.items, 1):
            pdf_bytes = _render_single_batch_pdf(
                value.customer,
                item,
            )

            drawing_name = _safe_pdf_name(
                item.drawing.drawing_no or f"drawing_{index}"
            )

            bundle.writestr(
                f"{index:02d}_{drawing_name}_quotation.pdf",
                pdf_bytes,
            )

    archive.seek(0)

    return StreamingResponse(
        archive,
        media_type="application/zip",
        headers={
            "Content-Disposition": 'attachment; filename="separate_quotations.zip"'
        },
    )
